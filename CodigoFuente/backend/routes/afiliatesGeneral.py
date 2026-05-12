"""
routers/affili.py
Router para gestión de lecturas de afiliados
"""

import csv
from decimal import Decimal
import io
from fastapi import APIRouter, Depends, HTTPException, logger, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import desc, func, extract
from typing import List, Optional, Dict
from datetime import date, datetime
from calendar import month_name

from models.lectura import Lectura
from models.meter import Medidor
from models.sector import Sector
from models.user import UsuarioSistema
from models.affiliate import UsuarioAfiliado
from db.session import SessionLocal
from security.jwt import verify_token

router = APIRouter(prefix="/afiliados", tags=["afiliados"])


def get_db():
    """Dependencia para obtener la sesión de base de datos"""
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def get_current_user(payload: dict, db: Session) -> UsuarioSistema:
    """Obtiene el usuario actual desde el payload del JWT"""
    user = db.query(UsuarioSistema).filter(
        UsuarioSistema.usuario == payload["sub"]
    ).first()
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario no encontrado"
        )
    
    return user


def get_current_afiliado(current_user: UsuarioSistema, db: Session) -> UsuarioAfiliado:
    """Obtiene el afiliado asociado al usuario actual"""
    afiliado = db.query(UsuarioAfiliado).filter(
        UsuarioAfiliado.id_usuario_sistema == current_user.id_usuario_sistema
    ).first()
    
    if not afiliado:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No se encontró información de afiliado para este usuario"
        )
    
    return afiliado


@router.get("/mis-lecturas/periodos-disponibles", response_model=dict)
def obtener_periodos_mis_lecturas(
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Obtiene años y meses disponibles agrupados por año
    Retorna: {anios_disponibles: [2026, 2025], periodos: {2026: [...], 2025: [...]}}
    """
    current_user = get_current_user(payload, db)

    afiliado = (
        db.query(UsuarioAfiliado.id_usuario_afi)
        .filter(
            UsuarioAfiliado.id_usuario_sistema == current_user.id_usuario_sistema
        )
        .first()
    )

    if not afiliado:
        return {
            "anios_disponibles": [],
            "periodos": {}
        }

    periodos = (
        db.query(
            extract('year', Lectura.fecha_lectura).label("anio"),
            extract('month', Lectura.fecha_lectura).label("mes"),
            func.count(Lectura.id_lectura).label("total_lecturas")
        )
        .join(Medidor, Medidor.id_medidor == Lectura.id_medidor)
        .filter(
            Medidor.id_usuario_afi == afiliado.id_usuario_afi,
            Lectura.activo == True
        )
        .group_by("anio", "mes")
        .order_by(
            extract('year', Lectura.fecha_lectura).desc(),
            extract('month', Lectura.fecha_lectura).desc()
        )
        .all()
    )

    # ✅ AGRUPAR POR AÑO
    periodos_agrupados = {}
    anios_disponibles = []

    for p in periodos:
        anio = int(p.anio)
        mes = int(p.mes)
        
        # Crear lista para el año si no existe
        if anio not in periodos_agrupados:
            periodos_agrupados[anio] = []
            anios_disponibles.append(anio)
        
        # Agregar mes con su nombre en español
        periodos_agrupados[anio].append({
            "mes": mes,
            "nombre_mes": obtener_nombre_mes(mes),
            "total_lecturas": p.total_lecturas
        })

    return {
        "anios_disponibles": anios_disponibles,
        "periodos": periodos_agrupados
    }

# ============================================
# IMPORTAR MODELO DE TARIFA
# ============================================
from models.tarifa import Tarifa

# ============================================
# ENDPOINT: OBTENER TARIFAS VIGENTES
# ============================================
@router.get("/tarifas-vigentes", response_model=Dict)
def obtener_tarifas_vigentes_endpoint(
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Obtiene las tarifas básica y de exceso vigentes
    Disponible para todos los usuarios autenticados
    """
    current_user = get_current_user(payload, db)
    
    tarifa_basica = db.query(Tarifa).filter(
        Tarifa.activo == True,
        Tarifa.es_vigente == True,
        Tarifa.tipo_tarifa == 'basico'
    ).first()
    
    tarifa_exceso = db.query(Tarifa).filter(
        Tarifa.activo == True,
        Tarifa.es_vigente == True,
        Tarifa.tipo_tarifa == 'exceso'
    ).first()
    
    return {
        "success": True,
        "tarifa_basica": {
            "id_tarifa": tarifa_basica.id_tarifa if tarifa_basica else None,
            "nombre": tarifa_basica.nombre if tarifa_basica else None,
            "limite_min_m3": float(tarifa_basica.limite_min_m3) if tarifa_basica else 0,
            "limite_max_m3": float(tarifa_basica.limite_max_m3) if tarifa_basica else 15,
            "precio_por_m3": float(tarifa_basica.precio_por_m3) if tarifa_basica else 0
        } if tarifa_basica else None,
        "tarifa_exceso": {
            "id_tarifa": tarifa_exceso.id_tarifa if tarifa_exceso else None,
            "nombre": tarifa_exceso.nombre if tarifa_exceso else None,
            "precio_por_m3": float(tarifa_exceso.precio_por_m3) if tarifa_exceso else 0
        } if tarifa_exceso else None
    }



@router.get("/mis-lecturas", response_model=List[dict])
def listar_mis_lecturas(
    anio: Optional[int] = Query(None, description="Año para filtrar (ej. 2025)"),
    mes: Optional[int] = Query(None, ge=1, le=12, description="Mes para filtrar (1-12)"),
    tipo_lectura: Optional[str] = Query(None, description="Tipo: reales, estimadas o todas"),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """Lista las lecturas del afiliado actual con clasificación de consumo"""
    
    current_user = get_current_user(payload, db)
    afiliado = db.query(UsuarioAfiliado)\
        .options(joinedload(UsuarioAfiliado.usuario_sistema))\
        .filter(UsuarioAfiliado.id_usuario_sistema == current_user.id_usuario_sistema)\
        .first()
    
    if not afiliado:
        return []
    
    # ✅ OBTENER TARIFAS VIGENTES
    tarifa_basica = db.query(Tarifa).filter(
        Tarifa.activo == True,
        Tarifa.es_vigente == True,
        Tarifa.tipo_tarifa == 'basico'
    ).first()
    
    limite_min = float(tarifa_basica.limite_min_m3) if tarifa_basica else 0
    limite_max = float(tarifa_basica.limite_max_m3) if tarifa_basica else 15
    
    # Query de lecturas (mantén tu código existente)
    query = db.query(
        Lectura.id_lectura,
        Lectura.id_medidor,
        Lectura.lectura_actual,
        Lectura.lectura_anterior,
        Lectura.consumo_m3,
        Lectura.fecha_lectura,
        Lectura.observacion,
        Lectura.es_estimada,
        Lectura.activo,
        Medidor.num_medidor,
        UsuarioAfiliado.cod_usuario_afi,
        UsuarioSistema.nombres.label('lector_nombres'),
        UsuarioSistema.apellidos.label('lector_apellidos'),
        Sector.nombre_sector,
    )\
    .join(Medidor, Medidor.id_medidor == Lectura.id_medidor)\
    .join(UsuarioAfiliado, UsuarioAfiliado.id_usuario_afi == Medidor.id_usuario_afi)\
    .outerjoin(UsuarioSistema, UsuarioSistema.id_usuario_sistema == Lectura.id_lector)\
    .outerjoin(Sector, Sector.id_sector == Medidor.id_sector)\
    .filter(
        Medidor.activo == True,
        Lectura.activo == True,
        Medidor.id_usuario_afi == afiliado.id_usuario_afi
    )
    
    # Aplicar filtros (mantén tu código existente)
    if anio and mes:
        fecha_inicio = date(anio, mes, 1)
        if mes == 12:
            fecha_fin = date(anio + 1, 1, 1)
        else:
            fecha_fin = date(anio, mes + 1, 1)
        query = query.filter(
            Lectura.fecha_lectura >= fecha_inicio,
            Lectura.fecha_lectura < fecha_fin
        )
    elif anio:
        fecha_inicio = date(anio, 1, 1)
        fecha_fin = date(anio + 1, 1, 1)
        query = query.filter(
            Lectura.fecha_lectura >= fecha_inicio,
            Lectura.fecha_lectura < fecha_fin
        )
    
    if tipo_lectura:
        tipo_lower = tipo_lectura.lower()
        if tipo_lower == 'reales':
            query = query.filter(Lectura.es_estimada == False)
        elif tipo_lower == 'estimadas':
            query = query.filter(Lectura.es_estimada == True)
    
    lecturas = query.order_by(Lectura.fecha_lectura.desc()).limit(200).all()
    
    # ✅ FORMATEAR RESPUESTA CON CLASIFICACIÓN
    def clasificar_consumo(consumo_m3):
        """Clasifica el consumo según tarifas"""
        if consumo_m3 < limite_min:
            return {
                "tipo": "bajo",
                "descripcion": "Bajo Mínimo",
                "color": "#3b82f6",  # Azul
                "icono": "arrow-down"
            }
        elif consumo_m3 <= limite_max:
            return {
                "tipo": "normal",
                "descripcion": "Rango Normal",
                "color": "#22c55e",  # Verde
                "icono": "check-circle"
            }
        else:
            exceso = consumo_m3 - limite_max
            return {
                "tipo": "exceso",
                "descripcion": f"Con Exceso (+{exceso:.2f} m³)",
                "color": "#ef4444",  # Rojo
                "icono": "alert-triangle"
            }
    
    nombre_afiliado = f"{afiliado.usuario_sistema.nombres} {afiliado.usuario_sistema.apellidos}".strip() if afiliado.usuario_sistema else "Sin nombre"
    codigo_afiliado = afiliado.cod_usuario_afi
    
    return [
        {
            "id_lectura": l.id_lectura,
            "id_medidor": l.id_medidor,
            "lectura_actual": l.lectura_actual,
            "lectura_anterior": l.lectura_anterior,
            "consumo_m3": l.consumo_m3,
            "fecha_lectura": l.fecha_lectura,
            "observacion": l.observacion,
            "es_estimada": l.es_estimada,
            "activo": l.activo,
            "anio": l.fecha_lectura.year if l.fecha_lectura else None,
            "mes": l.fecha_lectura.month if l.fecha_lectura else None,
            "nombre_mes": obtener_nombre_mes(l.fecha_lectura.month) if l.fecha_lectura else None,
            "medidor": {"num_medidor": l.num_medidor},
            "codigo_afiliado": codigo_afiliado,
            "nombre_afiliado": nombre_afiliado,
            "sector": l.nombre_sector or "Sin sector",
            "lector": {
                "nombres": l.lector_nombres,
                "apellidos": l.lector_apellidos
            },
            # ✅ NUEVA CLASIFICACIÓN
            "clasificacion_consumo": clasificar_consumo(l.consumo_m3 or 0)
        }
        for l in lecturas
    ]


@router.get("/consumo-por-periodo", response_model=Dict)
def obtener_consumo_por_periodo(
    anio: int = Query(...),
    mes: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    current_user = get_current_user(payload, db)
    afiliado = get_current_afiliado(current_user, db)

    # ===============================
    # RANGO DE FECHAS (CLAVE)
    # ===============================
    fecha_inicio = date(anio, mes, 1)
    fecha_fin = date(anio + 1, 1, 1) if mes == 12 else date(anio, mes + 1, 1)

    # ===============================
    # QUERY OPTIMIZADA
    # ===============================
    stats = (
        db.query(
            func.count(Lectura.id_lectura).label("total_lecturas"),
            func.coalesce(func.sum(Lectura.consumo_m3), 0).label("consumo_total"),
            func.coalesce(func.avg(Lectura.consumo_m3), 0).label("promedio_consumo")
        )
        .join(Medidor, Medidor.id_medidor == Lectura.id_medidor)
        .filter(
            Medidor.id_usuario_afi == afiliado.id_usuario_afi,
            Medidor.activo.is_(True),
            Lectura.activo.is_(True),
            Lectura.fecha_lectura >= fecha_inicio,
            Lectura.fecha_lectura < fecha_fin
        )
        .first()
    )

    return {
        "anio": anio,
        "mes": mes,
        "nombre_mes": obtener_nombre_mes(mes),
        "consumo_total": float(stats.consumo_total),
        "total_lecturas": stats.total_lecturas,
        "promedio_consumo": float(stats.promedio_consumo)
    }

@router.get("/estadisticas", response_model=Dict)
def obtener_estadisticas_generales(
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    current_user = get_current_user(payload, db)
    afiliado = get_current_afiliado(current_user, db)

    stats = (
        db.query(
            func.count(Lectura.id_lectura).label("total_lecturas"),
            func.coalesce(func.sum(Lectura.consumo_m3), 0).label("consumo_total"),
            func.coalesce(func.avg(Lectura.consumo_m3), 0).label("consumo_promedio"),
            func.count(func.distinct(Medidor.id_medidor)).label("total_medidores")
        )
        .join(Medidor, Medidor.id_medidor == Lectura.id_medidor)
        .filter(
            Medidor.id_usuario_afi == afiliado.id_usuario_afi,
            Medidor.activo.is_(True),
            Lectura.activo.is_(True)
        )
        .first()
    )

    return {
        "total_lecturas": stats.total_lecturas,
        "consumo_total": float(stats.consumo_total),
        "consumo_promedio": float(stats.consumo_promedio),
        "total_medidores": stats.total_medidores
    }


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

@router.get("/exportar-lecturas")
def exportar_lecturas_excel(
    anio: Optional[int] = Query(None),
    mes: Optional[int] = Query(None, ge=1, le=12),
    tipo_lectura: Optional[str] = Query(None),
    id_medidor: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Exporta las lecturas del afiliado a formato Excel (.xlsx)
    con encabezado informativo y diseño profesional
    """
    try:
        current_user = get_current_user(payload, db)
        afiliado = get_current_afiliado(current_user, db)

        # ===============================
        # CONSTRUCCIÓN DE FECHAS
        # ===============================
        fecha_inicio = fecha_fin = None
        if anio and mes:
            fecha_inicio = date(anio, mes, 1)
            fecha_fin = (
                date(anio + 1, 1, 1) if mes == 12 else date(anio, mes + 1, 1)
            )
        elif anio:
            fecha_inicio = date(anio, 1, 1)
            fecha_fin = date(anio + 1, 1, 1)

        # ===============================
        # QUERY OPTIMIZADA
        # ===============================
        query = (
            db.query(Lectura)
            .join(Medidor)
            .options(
                joinedload(Lectura.medidor)
                    .joinedload(Medidor.sector),
                joinedload(Lectura.medidor)
                    .joinedload(Medidor.usuario_afiliado)
                    .joinedload(UsuarioAfiliado.usuario_sistema),
                joinedload(Lectura.lector)
            )
            .filter(
                Medidor.id_usuario_afi == afiliado.id_usuario_afi,
                Medidor.activo.is_(True),
                Lectura.activo.is_(True)
            )
        )

        if id_medidor:
            query = query.filter(Medidor.id_medidor == id_medidor)

        if fecha_inicio and fecha_fin:
            query = query.filter(
                Lectura.fecha_lectura >= fecha_inicio,
                Lectura.fecha_lectura < fecha_fin
            )

        if tipo_lectura:
            tipo_lower = tipo_lectura.lower()
            if tipo_lower == "reales":
                query = query.filter(Lectura.es_estimada.is_(False))
            elif tipo_lower == "estimadas":
                query = query.filter(Lectura.es_estimada.is_(True))

        lecturas = query.order_by(Lectura.fecha_lectura.desc()).all()

        # ✅ Verificar si hay datos
        if not lecturas:
            raise HTTPException(
                status_code=404,
                detail="No se encontraron lecturas con los filtros especificados"
            )

        # ===============================
        # OBTENER DATOS DEL AFILIADO
        # ===============================
        usuario_afiliado = afiliado.usuario_sistema
        nombre_completo = f"{usuario_afiliado.nombres} {usuario_afiliado.apellidos}".strip() if usuario_afiliado else "Sin registro"
        codigo_afiliado = afiliado.cod_usuario_afi if afiliado else "N/A"
        
        # Obtener medidor y sector (del primer registro o filtrado)
        primer_lectura = lecturas[0]
        medidor_info = primer_lectura.medidor.num_medidor if primer_lectura.medidor else "N/A"
        sector_info = primer_lectura.medidor.sector.nombre_sector if primer_lectura.medidor and primer_lectura.medidor.sector else "Sin sector"

        # ===============================
        # CREAR LIBRO EXCEL
        # ===============================
        wb = Workbook()
        ws = wb.active
        ws.title = "Historial de Consumos"

        # ===============================
        # ESTILOS
        # ===============================
        # 🎨 Estilo para título principal
        title_font = Font(name='Calibri', size=16, bold=True, color='1F4E78')
        title_alignment = Alignment(horizontal='left', vertical='center')
        
        # 📋 Estilo para etiquetas del encabezado
        label_font = Font(name='Calibri', size=11, bold=True, color='1F4E78')
        value_font = Font(name='Calibri', size=11, color='000000')
        info_alignment = Alignment(horizontal='left', vertical='center')
        
        # 🎨 Estilo para encabezados de tabla (AZUL CLARO)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 📐 Bordes
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # 📝 Estilo para datos
        data_alignment_left = Alignment(horizontal='left', vertical='center')
        data_alignment_center = Alignment(horizontal='center', vertical='center')
        data_alignment_right = Alignment(horizontal='right', vertical='center')

        # ===============================
        # ENCABEZADO INFORMATIVO
        # ===============================
        current_row = 1
        
        # Título principal
        ws.cell(row=current_row, column=1, value="HISTORIAL DE CONSUMOS")
        ws.cell(row=current_row, column=1).font = title_font
        ws.cell(row=current_row, column=1).alignment = title_alignment
        current_row += 2
        
        # Información del afiliado
        info_data = [
            ("Nombres:", nombre_completo),
            ("Código Afiliado:", codigo_afiliado),
            ("Medidor:", medidor_info),
            ("Sector:", sector_info),
            ("Fecha de Exportación:", datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        ]
        
        for label, value in info_data:
            ws.cell(row=current_row, column=1, value=label).font = label_font
            ws.cell(row=current_row, column=1).alignment = info_alignment
            ws.cell(row=current_row, column=2, value=value).font = value_font
            ws.cell(row=current_row, column=2).alignment = info_alignment
            current_row += 1
        
        current_row += 1  # Espacio antes de la tabla

        # ===============================
        # ENCABEZADOS DE TABLA
        # ===============================
        headers = [
            'Fecha Lectura',
            'Año',
            'Mes',
            'Lectura Anterior (m³)',
            'Lectura Actual (m³)',
            'Consumo (m³)',
            'Tipo',
            'Lector',
            'Observación',
            'Estado'
        ]
        
        header_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        current_row += 1

        # ===============================
        # DATOS
        # ===============================
        for lectura in lecturas:
            medidor = lectura.medidor
            lector = lectura.lector
            
            row_data = [
                lectura.fecha_lectura.strftime('%d/%m/%Y') if lectura.fecha_lectura else '',
                lectura.fecha_lectura.year if lectura.fecha_lectura else '',
                obtener_nombre_mes(lectura.fecha_lectura.month) if lectura.fecha_lectura else '',
                lectura.lectura_anterior if lectura.lectura_anterior is not None else 0,
                lectura.lectura_actual if lectura.lectura_actual is not None else 0,
                lectura.consumo_m3 if lectura.consumo_m3 is not None else 0,
                'Estimada' if lectura.es_estimada else 'Real',
                f"{lector.nombres} {lector.apellidos}".strip() if lector else 'No registrado',
                (lectura.observacion or '').replace('\n', ' ').replace('\r', ''),
                'Activo' if lectura.activo else 'Inactivo'
            ]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_num, value=value)
            
            current_row += 1

        # ===============================
        # FORMATO DE CELDAS DE DATOS
        # ===============================
        data_start_row = header_row + 1
        for row_idx in range(data_start_row, current_row):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
                # Alineación según columna
                if col_idx in [1, 2, 3, 7, 10]:  # Fecha, Año, Mes, Tipo, Estado
                    cell.alignment = data_alignment_center
                elif col_idx in [4, 5, 6]:  # Valores numéricos
                    cell.alignment = data_alignment_right
                    cell.number_format = '#,##0.00'  # Formato numérico con 2 decimales
                else:
                    cell.alignment = data_alignment_left

        # ===============================
        # AJUSTE AUTOMÁTICO DE ANCHOS
        # ===============================
        column_widths = {
            'A': 14,  # Fecha
            'B': 8,   # Año
            'C': 12,  # Mes
            'D': 18,  # Lectura Ant
            'E': 18,  # Lectura Act
            'F': 14,  # Consumo
            'G': 12,  # Tipo
            'H': 25,  # Lector
            'I': 35,  # Observación
            'J': 10   # Estado
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # ===============================
        # CARACTERÍSTICAS ADICIONALES
        # ===============================
        # ✅ Congelar panel de encabezados de tabla
        ws.freeze_panes = f'A{header_row + 1}'
        
        # ✅ Aplicar filtros automáticos solo en la tabla
        ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(headers))}{current_row - 1}'

        # ===============================
        # NOMBRE DINÁMICO DEL ARCHIVO
        # ===============================
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        periodo = ""
        if anio and mes:
            periodo = f"_{obtener_nombre_mes(mes)}_{anio}"
        elif anio:
            periodo = f"_{anio}"
        
        filename = f"historial_consumos{periodo}_{timestamp}.xlsx"

        # ===============================
        # GUARDAR EN MEMORIA Y RETORNAR
        # ===============================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Cache-Control": "no-cache"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exportando lecturas Excel: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail="Error generando el archivo de exportación"
        )



@router.get("/mis-medidores")
def obtener_mis_medidores(
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    # Una sola query con todos los JOINs necesarios — sin lazy-loads
    from sqlalchemy import func as sqlfunc
    
    row = (
        db.query(
            UsuarioAfiliado.id_usuario_afi,
            UsuarioAfiliado.cod_usuario_afi,
            UsuarioAfiliado.fecha_afiliacion,
            UsuarioAfiliado.activo,
            UsuarioSistema.id_usuario_sistema,
            UsuarioSistema.nombres,
            UsuarioSistema.apellidos,
            UsuarioSistema.cedula,
            UsuarioSistema.email,
            UsuarioSistema.telefono,
            UsuarioSistema.direccion,
            Sector.id_sector.label("sector_afi_id"),
            Sector.nombre_sector.label("sector_afi_nombre"),
        )
        .join(UsuarioSistema, UsuarioAfiliado.id_usuario_sistema == UsuarioSistema.id_usuario_sistema)
        .outerjoin(Sector, UsuarioAfiliado.id_sector == Sector.id_sector)
        .filter(
            UsuarioSistema.usuario == payload["sub"],
            UsuarioAfiliado.activo == True
        )
        .first()
    )

    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No tienes un perfil de afiliado registrado"
        )

    # Medidores con su sector en UNA sola query
    medidores_rows = (
        db.query(
            Medidor.id_medidor,
            Medidor.num_medidor,
            Medidor.activo,
            Medidor.latitud,
            Medidor.longitud,
            Medidor.altitud,
            Sector.id_sector.label("sector_id"),
            Sector.nombre_sector.label("sector_nombre"),
        )
        .outerjoin(Sector, Medidor.id_sector == Sector.id_sector)
        .filter(Medidor.id_usuario_afi == row.id_usuario_afi)
        .all()
    )

    medidores_list = [
        {
            "id_medidor": m.id_medidor,
            "num_medidor": m.num_medidor,
            "activo": m.activo,
            "latitud": float(m.latitud) if m.latitud else None,
            "longitud": float(m.longitud) if m.longitud else None,
            "altitud": float(m.altitud) if m.altitud else None,
            "sector": {
                "id_sector": m.sector_id,
                "nombre_sector": m.sector_nombre,
                "activo": True
            } if m.sector_id else None
        }
        for m in medidores_rows
    ]

    return {
        "medidores": medidores_list,
        "total_medidores": len(medidores_list),
        "afiliado": {
            "id_usuario_afi": row.id_usuario_afi,
            "cod_usuario_afi": row.cod_usuario_afi,
            "fecha_afiliacion": row.fecha_afiliacion.strftime("%Y-%m-%d") if row.fecha_afiliacion else None,
            "activo": row.activo,
            "sector": {
                "id_sector": row.sector_afi_id,
                "nombre_sector": row.sector_afi_nombre
            } if row.sector_afi_id else None,
            "usuario_sistema": {
                "id_usuario_sistema": row.id_usuario_sistema,
                "nombres": row.nombres,
                "apellidos": row.apellidos,
                "cedula": row.cedula,
                "email": row.email,
                "telefono": row.telefono,
                "direccion": row.direccion
            }
        }
    }


def obtener_nombre_mes(mes: int) -> str:
    """Retorna el nombre del mes en español"""
    meses = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    return meses.get(mes, f"Mes {mes}")
