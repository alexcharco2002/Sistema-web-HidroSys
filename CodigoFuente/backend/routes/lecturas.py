from decimal import Decimal
from fastapi import APIRouter, Body, Depends, Form, HTTPException, status, Query
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from models.factura import Factura
from models.sector import Sector
from models.tarifa import Tarifa
from schemas.tarifa import TarifaResponse
from sqlalchemy.orm import Session, aliased
from sqlalchemy.exc import IntegrityError
from sqlalchemy import and_, extract, func
from typing import List, Optional
from datetime import date, datetime
import io
from fastapi.responses import StreamingResponse
from fastapi import UploadFile, File
from calendar import month_name
import locale
from collections import defaultdict

from openpyxl.styles import Protection  # Importar Protection para proteger/desproteger celdas
from models.lectura import Lectura
from models.meter import Medidor
from models.user import UsuarioSistema
from models.role import RolAccion
from models.affiliate import UsuarioAfiliado  # Importar modelo de UsuarioAfiliado
from schemas.lectura import (
    LecturaCreate,
    LecturaUpdate,
    LecturaResponse,
    LecturaStats,
    LecturaBulkCreate,
    LecturaBulkCreateRequest,
    LecturaBulkResponse,
    LecturaBulkResult,
    LecturaBulkError
)
from utils.notifications import registrar_notificacion
from utils.audit_logger import registrar_auditoria
from db.session import SessionLocal
from security.jwt import verify_token

from utils.facturacion import calcular_descuento, generar_factura_desde_lectura, reactivar_factura_anulada, recalcular_factura, regenerar_factura_desde_lectura_anulada # para generar facturas automaticas 


router = APIRouter(prefix="/lecturas", tags=["lecturas"])

from typing import Optional
from datetime import date

def get_db():
    """Dependencia para obtener la sesión de base de datos"""
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ============================================================================
# HELPER: Obtener usuario actual desde el token
# ============================================================================

def get_current_user(payload: dict, db: Session) -> UsuarioSistema:
    """Obtiene el usuario actual desde el payload del JWT"""
    user = db.query(UsuarioSistema).filter(
        UsuarioSistema.usuario == payload["sub"]
    ).first()
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario no encontrado"
        )
    
    return user


# ============================================================================
# HELPER: Verificar permisos de usuario
# ============================================================================

def check_permission(user: UsuarioSistema, db: Session, module: str, action: str = None) -> bool:
    """
    Verifica si el usuario tiene permiso para una acción.
    """
    module = module.lower().strip()
    action = action.lower().strip() if action else None
    
    permisos = db.query(RolAccion).filter(
        RolAccion.id_rol == user.id_rol,
        RolAccion.activo == True
    ).all()
    
    acciones_usuario = set()
    
    for permiso in permisos:
        if not permiso.nombre_accion:
            continue
        
        perm_module = permiso.nombre_accion.lower().strip()
        perm_action = (permiso.tipo_accion or '').lower().strip()
        
        if perm_module != module:
            continue
        
        if perm_action in ['crud', 'operaciones crud']:
            return True
        
        acciones_usuario.add(perm_action)
    
    if action is None:
        return bool(acciones_usuario)
    
    if action in ['leer', 'lectura']:
        if any(a in acciones_usuario for a in ['lectura', 'leer', 'crear', 'actualizar', 'eliminar']):
            return True
    
    return action in acciones_usuario


def require_permission(user: UsuarioSistema, db: Session, module: str, action: str = None):
    """Verifica permiso y lanza excepción si no lo tiene"""
    if not check_permission(user, db, module, action):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"No tienes permisos para {action or 'acceder a'} {module}"
        )

def require_any_permission(
    user: UsuarioSistema,
    db: Session,
    permissions: list[tuple[str, str | None]]
):
    """
    Permite acceso si el usuario tiene AL MENOS uno de los permisos indicados
    """
    for module, action in permissions:
        if check_permission(user, db, module, action):
            return
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="No tienes permisos para acceder a este recurso"
    )


def construir_periodo_consumo(anio: int, mes: int) -> str:
    return f"{anio}-{mes:02d}"


def periodo_desde_fecha(fecha: date) -> str:
    return construir_periodo_consumo(fecha.year, fecha.month)


def periodo_desde_mes_anio(periodo: str) -> tuple[int, int]:
    """Convierte un periodo YYYY-MM a (anio, mes)."""
    anio, mes = periodo.split("-")
    return int(anio), int(mes)


# ============================================================================
# HELPER: Convertir lectura a respuesta con información completa
# ============================================================================

def lectura_to_response(lectura: Lectura) -> dict:
    """Convierte una lectura con información del medidor y lector"""
    medidor = lectura.medidor
    lector = lectura.lector
    
    return {
        "id_lectura": lectura.id_lectura,
        "id_medidor": lectura.id_medidor,
        "lectura_actual": lectura.lectura_actual,
        "lectura_anterior": lectura.lectura_anterior,
        "consumo_m3": lectura.consumo_m3,
        "fecha_lectura": lectura.fecha_lectura.isoformat() if lectura.fecha_lectura else None,
        "periodo_consumo": lectura.periodo_consumo,
        "id_lector": lectura.id_lector,
        "observacion": lectura.observacion,
        "activo": lectura.activo,
        "es_estimada": lectura.es_estimada,
        "medidor": {
            "id_medidor": medidor.id_medidor,
            "num_medidor": medidor.num_medidor
        } if medidor else None,
        "lector": {
            "id_usuario_sistema": lector.id_usuario_sistema,
            "nombres": lector.nombres,
            "apellidos": lector.apellidos
        } if lector else None
    }

# ========================================
# ENDPOINT PARA LISTAR AFILIADOS CON MEDIDORES (OPTIMIZADO)
# ========================================

@router.get("/medidores/lista/completa", response_model=dict)
def listar_afiliados_con_medidores(
    mes: Optional[int] = Query(None, ge=1, le=12, description="Mes del periodo para filtrar"),
    anio: Optional[int] = Query(None, ge=2020, description="Año del periodo para filtrar"),
    incluir_con_lectura: bool = Query(False, description="Incluir medidores con lectura en el periodo"),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Lista afiliados con sus medidores para registrar lecturas.
    
    Parámetros:
    - mes, anio: Filtran medidores sin lectura en ese periodo
    - incluir_con_lectura: Si es True, incluye todos los medidores
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "lectura")

    try:
        print(f"\n{'='*60}")
        print(f"📋 LISTANDO AFILIADOS CON MEDIDORES")
        print(f"{'='*60}")
        
        # ============================================
        # PASO 1: OBTENER MEDIDORES CON LECTURA EN EL PERIODO
        # ============================================
        medidores_con_lectura_ids = set()
        
        if mes and anio and not incluir_con_lectura:
            periodo = construir_periodo_consumo(anio, mes)
            lecturas_periodo = db.query(Lectura.id_medidor).filter(
                Lectura.periodo_consumo == periodo
            ).distinct().all()
            
            medidores_con_lectura_ids = {lectura[0] for lectura in lecturas_periodo}
            print(f"📅 Periodo: {mes:02d}/{anio}")
            print(f"⚠️ Medidores con lectura en periodo: {len(medidores_con_lectura_ids)}")
        
        # ============================================
        # PASO 2: SUBCONSULTA PARA ÚLTIMA LECTURA
        # ============================================
        subq_ultima_lectura = (
            db.query(
                Lectura.id_medidor,
                func.max(Lectura.periodo_consumo).label("max_periodo")
            )
            .filter(Lectura.activo == True)
            .group_by(Lectura.id_medidor)
            .subquery()
        )

        # ============================================
        # PASO 3: CONSULTA PRINCIPAL OPTIMIZADA
        # ============================================
        query = (
            db.query(
                Medidor.id_medidor,
                Medidor.num_medidor,
                Medidor.activo,
                UsuarioAfiliado.id_usuario_afi,
                UsuarioAfiliado.cod_usuario_afi,
                UsuarioSistema.nombres,
                UsuarioSistema.apellidos,
                UsuarioSistema.cedula,
                Sector.nombre_sector,
                Lectura.lectura_actual.label("lectura_anterior"),
                Lectura.fecha_lectura.label("fecha_ultima_lectura"),
                Lectura.periodo_consumo.label("periodo_ultima_lectura")
            )
            .join(
                UsuarioAfiliado,
                Medidor.id_usuario_afi == UsuarioAfiliado.id_usuario_afi
            )
            .join(
                UsuarioSistema,
                UsuarioAfiliado.id_usuario_sistema == UsuarioSistema.id_usuario_sistema
            )
            .outerjoin(
                Sector,
                func.coalesce(Medidor.id_sector, UsuarioAfiliado.id_sector) == Sector.id_sector
            )
            .outerjoin(
                subq_ultima_lectura,
                subq_ultima_lectura.c.id_medidor == Medidor.id_medidor
            )
            .outerjoin(
                Lectura,
                and_(
                    Lectura.id_medidor == Medidor.id_medidor,
                    Lectura.periodo_consumo == subq_ultima_lectura.c.max_periodo,
                    Lectura.activo == True
                )
            )
            .filter(
                Medidor.activo == True,
                UsuarioAfiliado.activo == True
            )
        )
        
        # ============================================
        # PASO 4: EXCLUIR MEDIDORES CON LECTURA EN PERIODO
        # ============================================
        if medidores_con_lectura_ids:
            query = query.filter(
                Medidor.id_medidor.notin_(medidores_con_lectura_ids)
            )
            print(f"🔍 Filtrando medidores sin lectura en {mes:02d}/{anio}")
        
        # Ordenar y ejecutar
        resultados = query.order_by(UsuarioAfiliado.cod_usuario_afi).all()
        
        print(f"✅ Afiliados encontrados: {len(resultados)}")
        print(f"{'='*60}\n")

        if not resultados:
            return {
                "afiliados": [],
                "total": 0,
                "periodo": {
                    "mes": mes,
                    "anio": anio,
                    "filtrado": bool(mes and anio)
                },
                "mensaje": f"Lecturas Completadas - No hay medidores disponibles para lecturas del periodo {mes:02d}/{anio}" if mes and anio else "No hay medidores activos"
            }

        # ============================================
        # PASO 5: TRANSFORMAR A LISTA
        # ============================================
        afiliados = []
        for r in resultados:
            afiliado = {
                "id_medidor": r.id_medidor,
                "num_medidor": r.num_medidor,
                "activo": r.activo,
                "id_usuario_afi": r.id_usuario_afi,
                "cod_usuario_afi": r.cod_usuario_afi,
                "nombre_completo": f"{r.nombres or ''} {r.apellidos or ''}".strip(),
                "nombres": r.nombres,
                "apellidos": r.apellidos,
                "cedula": r.cedula,
                "sector": r.nombre_sector or "Sin sector",
                "nombre_sector": r.nombre_sector or "Sin sector",
                "lectura_anterior": float(r.lectura_anterior) if r.lectura_anterior else 0,
                "fecha_ultima_lectura": r.fecha_ultima_lectura.strftime('%Y-%m-%d') if r.fecha_ultima_lectura else None,
                "periodo_ultima_lectura": r.periodo_ultima_lectura,
                "tiene_lectura_anterior": r.lectura_anterior is not None
            }
            afiliados.append(afiliado)

        return {
            "success": True,
            "afiliados": afiliados,
            "total": len(afiliados),
            "periodo": {
                "mes": mes,
                "anio": anio,
                "filtrado": bool(mes and anio),
                "excluidos": len(medidores_con_lectura_ids)
            }
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener lista de afiliados: {str(e)}"
        )


@router.get("", response_model=List[dict])
def listar_lecturas_optimizado(
    mes: int = Query(None, ge=1, le=12, description="Mes del periodo (1-12)"),
    anio: int = Query(None, ge=2020, description="Año del periodo"),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    current_user = get_current_user(payload, db)
    require_any_permission(
        current_user,
        db,
        [
            ("lecturas", "lectura"),
            ("lecturas", "crud"),
            ("historialconsumo", "crud"),
        ]
    )

    try:
        # 🔥 Crear DOS alias diferentes para UsuarioSistema
        # Uno para el Afiliado y otro para el Lector
        UsuarioAfiliado_Alias = aliased(UsuarioSistema, name="usuario_afiliado")
        UsuarioLector_Alias = aliased(UsuarioSistema, name="usuario_lector")
        
        # 🔹 Query base
        query = (
            db.query(
                Lectura.id_lectura,
                Lectura.id_medidor,  # 🔥 AGREGADO: id_medidor
                Lectura.fecha_lectura,
                Lectura.periodo_consumo,
                Lectura.lectura_actual,
                Lectura.lectura_anterior,
                Lectura.consumo_m3,
                Lectura.observacion,
                Lectura.activo,
                Lectura.es_estimada,
                
                # Datos del medidor
                Medidor.num_medidor,
                
                # Datos del afiliado (dueño del medidor)
                UsuarioAfiliado.cod_usuario_afi.label("codigo_afiliado"),
                UsuarioAfiliado_Alias.nombres.label("afiliado_nombres"),
                UsuarioAfiliado_Alias.apellidos.label("afiliado_apellidos"),
                
                # Datos del sector
                Sector.nombre_sector,
                
                # 🔥 Datos del lector (quien tomó la lectura)
                UsuarioLector_Alias.nombres.label("lector_nombres"),
                UsuarioLector_Alias.apellidos.label("lector_apellidos"),
            )
            # Join con Medidor
            .join(Medidor, Lectura.id_medidor == Medidor.id_medidor)
            
            # Join con UsuarioAfiliado (tabla intermedia)
            .join(UsuarioAfiliado, Medidor.id_usuario_afi == UsuarioAfiliado.id_usuario_afi)
            
            # 🔥 Join con UsuarioSistema (datos del AFILIADO)
            .join(
                UsuarioAfiliado_Alias,
                UsuarioAfiliado.id_usuario_sistema == UsuarioAfiliado_Alias.id_usuario_sistema
            )
            
            # Join con Sector (opcional)
            .outerjoin(Sector, Sector.id_sector == Medidor.id_sector)
            
            # 🔥 Join con UsuarioSistema (datos del LECTOR) - OPCIONAL
            .outerjoin(
                UsuarioLector_Alias,
                Lectura.id_lector == UsuarioLector_Alias.id_usuario_sistema
            )
        )

        # 🔥 Filtrar por periodo si se proporciona
        if mes is not None and anio is not None:
            periodo = construir_periodo_consumo(anio, mes)
            query = query.filter(
                Lectura.periodo_consumo == periodo
            )
        
        # Ordenar por periodo de consumo; fecha_lectura queda como desempate.
        lecturas = query.order_by(
            Lectura.periodo_consumo.desc(),
            Lectura.fecha_lectura.desc()
        ).limit(500).all()

        print(f"✅ Total lecturas encontradas: {len(lecturas)}")

        return [
            {
                "id_lectura": l.id_lectura,
                "id_medidor": l.id_medidor,  # 🔥 AGREGADO en la respuesta
                "fecha_lectura": l.fecha_lectura,
                "periodo_consumo": l.periodo_consumo,
                "lectura_actual": l.lectura_actual,
                "lectura_anterior": l.lectura_anterior,
                "consumo_m3": l.consumo_m3,
                "observacion": l.observacion,
                "activo": l.activo,
                "es_estimada": l.es_estimada,
                
                # Medidor
                "num_medidor": l.num_medidor,
                
                # Afiliado (dueño del medidor)
                "codigo_afiliado": l.codigo_afiliado,
                "nombre_afiliado": f"{l.afiliado_nombres} {l.afiliado_apellidos}",
                
                # Sector
                "sector": l.nombre_sector or "Sin sector",
                
                # 🔥 Lector (quien tomó la lectura)
                "lector_nombre": (
                    f"{l.lector_nombres} {l.lector_apellidos}".strip() 
                    if l.lector_nombres else "No registrado"
                ),
            }
            for l in lecturas
        ]

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al listar lecturas: {str(e)}"
        )


@router.get("/stats/count", response_model=LecturaStats)
def obtener_estadisticas_lecturas(
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Obtiene estadísticas de lecturas
    Requiere permiso: lecturas.lectura o lecturas.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "lectura")
    
    total = db.query(Lectura).count()
    activos = db.query(Lectura).filter(Lectura.activo == True).count()
    inactivos = db.query(Lectura).filter(Lectura.activo == False).count()
    
    # Consumo total
    consumo_total = db.query(db.func.sum(Lectura.consumo_m3)).scalar() or 0
    
    return {
        "total": total,
        "activos": activos,
        "inactivos": inactivos,
        "consumo_total": consumo_total
    }


@router.get("/{id_lectura}", response_model=dict)
def obtener_lectura(
    id_lectura: int,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Obtiene una lectura específica por ID
    Requiere permiso: lecturas.lectura o lecturas.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "lectura")
    
    
    lectura = db.query(Lectura).filter(Lectura.id_lectura == id_lectura).first()
    
    if not lectura:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lectura no encontrada"
        )
    
    return lectura_to_response(lectura)

@router.post("", response_model=dict, status_code=status.HTTP_201_CREATED)
def crear_lectura(
    lectura_data: LecturaCreate,
    generar_factura: bool = Query(True, description="Generar factura automáticamente"),
    tipo_descuento: str = Query('ninguno', description="Tipo: ninguno/porcentaje/valor"),
    valor_descuento: float = Query(0.0, ge=0, description="Valor del descuento"),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Crea una nueva lectura.
    La tarifa se determina automáticamente según el consumo.
    Opcionalmente genera la factura automáticamente.
    Requiere permiso: lecturas.crear o lecturas.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "crear")
    
    # Verificar que el medidor existe
    medidor = db.query(Medidor).filter(
        Medidor.id_medidor == lectura_data.id_medidor
    ).first()
    
    if not medidor:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Medidor no encontrado"
        )
    
    # Validación: evitar doble lectura en el mismo mes
    periodo_consumo = lectura_data.periodo_consumo or periodo_desde_fecha(lectura_data.fecha_lectura)

    lectura_mes_existente = db.query(Lectura).filter(
        Lectura.id_medidor == lectura_data.id_medidor,
        Lectura.periodo_consumo == periodo_consumo,
        Lectura.activo == True
    ).first()

    if lectura_mes_existente:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Ya existe una lectura registrada para este medidor "
                f"en el periodo de consumo {periodo_consumo}."
            )
        )
    
    # Obtener información del afiliado
    afiliado = medidor.usuario_afiliado if medidor else None
    usuario_afiliado = afiliado.usuario_sistema if afiliado else None
    
    if usuario_afiliado:
        nombre_afiliado = f"{usuario_afiliado.nombres} {usuario_afiliado.apellidos}"
        id_usuario_afiliado = usuario_afiliado.id_usuario_sistema
    else:
        nombre_afiliado = "Usuario desconocido"
        id_usuario_afiliado = None
    
    # Crear nueva lectura
    nueva_lectura = Lectura(
        id_medidor=lectura_data.id_medidor,
        lectura_actual=lectura_data.lectura_actual,
        lectura_anterior=lectura_data.lectura_anterior,
        consumo_m3=lectura_data.consumo_m3,
        fecha_lectura=lectura_data.fecha_lectura,
        periodo_consumo=periodo_consumo,
        id_lector=current_user.id_usuario_sistema,
        observacion=lectura_data.observacion,
        activo=lectura_data.activo,
        es_estimada=lectura_data.es_estimada
    )
    
    try:
        db.add(nueva_lectura)
        db.flush()

        lectura_id = nueva_lectura.id_lectura

        # Generar factura
        factura_generada = None
        mensaje_factura = ""
        
        if generar_factura:
            exito, mensaje, factura_generada = generar_factura_desde_lectura(
                db=db,
                lectura=nueva_lectura,
                tipo_descuento=tipo_descuento,
                valor_descuento=valor_descuento,
                aplicar_servicios=True,
                aplicar_multas=True
            )
            
            if exito:
                mensaje_factura = "Factura generada correctamente."
            else:
                mensaje_factura = f"No se pudo generar la factura: {mensaje}"
        
        # Commit final
        db.commit()
        db.refresh(nueva_lectura)
        
        # Auditoría
        registrar_auditoria(
            db=db,
            accion="CREATE",
            descripcion=f"Lectura creada para medidor {medidor.num_medidor} (Consumo: {nueva_lectura.consumo_m3}m³) - {mensaje_factura}",
            id_usuario=current_user.id_usuario_sistema
        )
        
        # Notificación para el lector
        registrar_notificacion(
            db=db,
            id_usuario=current_user.id_usuario_sistema,
            titulo="Lectura creada",
            mensaje=f"Lectura registrada para el medidor {medidor.num_medidor}. Consumo: {nueva_lectura.consumo_m3} m3. {mensaje_factura}",
            tipo="exito"
        )
        
        # Notificación para el afiliado
        if id_usuario_afiliado:
            mensaje_afiliado = f"Se registró una lectura de {nueva_lectura.consumo_m3} m3 para tu medidor {medidor.num_medidor}."
            
            if factura_generada:
                mensaje_afiliado += f" Factura {factura_generada.num_factura} generada por ${factura_generada.total}."
            
            registrar_notificacion(
                db=db,
                id_usuario=id_usuario_afiliado,
                titulo="Nueva lectura registrada",
                mensaje=mensaje_afiliado,
                tipo="info"
            )
        
        # Preparar respuesta
        response_data = lectura_to_response(nueva_lectura)

        if factura_generada:
            # Obtener tarifa aplicada
            tarifa_aplicada = db.query(Tarifa).filter(
                Tarifa.id_tarifa == factura_generada.id_tarifa
            ).first()
            
            response_data['factura_generada'] = {
                'id_factura': factura_generada.id_factura,
                'num_factura': factura_generada.num_factura,
                'total': float(factura_generada.total),
                'tarifa_aplicada': tarifa_aplicada.tipo_tarifa if tarifa_aplicada else "N/A",
                'periodo': factura_generada.periodo,
                'mensaje': mensaje_factura
            }
        else:
            response_data['factura_generada'] = None
            response_data['mensaje_factura'] = mensaje_factura
        
        return response_data
    
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear la lectura: {str(e)}"
        )



@router.put("/{id_lectura}", response_model=dict)
def actualizar_lectura(
    id_lectura: int,
    lectura_data: LecturaUpdate,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Actualiza una lectura existente y gestiona factura según estado:
    - PENDIENTE: Recalcula factura existente
    - ANULADA: Reactiva y recalcula la factura (mismo número)
    - PAGADA: No permite actualización
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "actualizar")
    
    # Buscar la lectura
    lectura = db.query(Lectura).filter(Lectura.id_lectura == id_lectura).first()
    
    if not lectura:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lectura no encontrada"
        )

    # ============================================
    # VERIFICAR FACTURA RELACIONADA
    # ============================================
    factura_relacionada = db.query(Factura).filter(
        Factura.id_lectura == id_lectura
    ).first()
    
    # ⚠️ BLOQUEAR si factura está PAGADA
    if factura_relacionada and factura_relacionada.estado_factura == 'pagada':
        return {
            "success": False,
            "accion": "no_actualizado",
            "message": "No se puede actualizar la lectura porque tiene una factura pagada.",
            "info": {
                "id_factura": factura_relacionada.id_factura,
                "numero_factura": factura_relacionada.num_factura,
                "estado": factura_relacionada.estado_factura
            }
        }

    # ============================================
    # ACTUALIZAR LECTURA
    # ============================================
    update_data = lectura_data.model_dump(exclude_unset=True)
    
    # Validar duplicados
    if "periodo_consumo" in update_data or "id_medidor" in update_data:
        nuevo_periodo = update_data.get("periodo_consumo", lectura.periodo_consumo)
        nuevo_medidor = update_data.get("id_medidor", lectura.id_medidor)

        duplicado = db.query(Lectura).filter(
            Lectura.id_medidor == nuevo_medidor,
            Lectura.periodo_consumo == nuevo_periodo,
            Lectura.id_lectura != id_lectura,
            Lectura.activo == True
        ).first()

        if duplicado:
            return {
                "success": False,
                "accion": "no_actualizado",
                "message": "Ya existe otra lectura para ese medidor en ese mes.",
                "info": {
                    "id_lectura_existente": duplicado.id_lectura,
                    "periodo_consumo": nuevo_periodo
                }
            }
    
    # Aplicar cambios a la lectura
    for key, value in update_data.items():
        setattr(lectura, key, value)
    
    # Recalcular consumo si cambiaron las lecturas
    if "lectura_actual" in update_data or "lectura_anterior" in update_data:
        lectura.consumo_m3 = lectura.lectura_actual - lectura.lectura_anterior
    
    try:
        db.commit()
        db.refresh(lectura)
        
        # ============================================
        # GESTIÓN DE FACTURA SEGÚN ESTADO
        # ============================================
        accion_factura = "sin_factura"
        info_factura = {}
        mensaje_adicional = ""
        
        if factura_relacionada:
            if factura_relacionada.estado_factura == 'pendiente':
                # ✅ CASO 1: RECALCULAR factura pendiente
                print(f"\n🔄 Factura PENDIENTE: recalculando...")
                
                try:
                    factura_actualizada = recalcular_factura(db, factura_relacionada, lectura)
                    
                    accion_factura = "factura_recalculada"
                    info_factura = {
                        "id_factura": factura_actualizada.id_factura,
                        "numero_factura": factura_actualizada.num_factura,
                        "consumo_m3": factura_actualizada.consumo_m3,
                        "exceso_m3": factura_actualizada.exceso_m3,
                        "nuevo_total": float(factura_actualizada.total),
                        "estado": factura_actualizada.estado_factura
                    }
                    mensaje_adicional = "Factura recalculada (mantiene multas y servicios)"
                    
                except Exception as e:
                    print(f"❌ Error recalculando: {e}")
                    return {
                        "success": False,
                        "accion": "error_recalculo",
                        "message": f"Lectura actualizada pero error al recalcular factura: {str(e)}"
                    }
                
            elif factura_relacionada.estado_factura == 'anulada':
                # ✅ CASO 2: REACTIVAR factura anulada (MISMO NÚMERO)
                print(f"\n♻️ Factura ANULADA: reactivando...")
                
                try:
                    factura_reactivada = reactivar_factura_anulada(db, factura_relacionada, lectura)
                    
                    accion_factura = "factura_reactivada"
                    info_factura = {
                        "id_factura": factura_reactivada.id_factura,
                        "numero_factura": factura_reactivada.num_factura,
                        "consumo_m3": factura_reactivada.consumo_m3,
                        "nuevo_total": float(factura_reactivada.total),
                        "estado": factura_reactivada.estado_factura,
                        "estado_anterior": "anulada"
                    }
                    mensaje_adicional = "Factura reactivada a estado pendiente"
                    
                except Exception as e:
                    print(f"❌ Error reactivando: {e}")
                    return {
                        "success": False,
                        "accion": "error_reactivacion",
                        "message": f"Lectura actualizada pero error al reactivar factura: {str(e)}"
                    }
        
        # ============================================
        # AUDITORÍA Y NOTIFICACIÓN
        # ============================================
        registrar_auditoria(
            db=db,
            accion="UPDATE",
            descripcion=f"Lectura {id_lectura} actualizada - {accion_factura}",
            id_usuario=current_user.id_usuario_sistema
        )
        
        mensaje_notif = f"Lectura modificada. {mensaje_adicional if mensaje_adicional else 'Sin factura asociada'}"
        registrar_notificacion(
            db=db,
            id_usuario=current_user.id_usuario_sistema,
            titulo="Lectura actualizada",
            mensaje=mensaje_notif,
            tipo="info"
        )
        
        return {
            "success": True,
            "accion": "actualizado",
            "data": lectura_to_response(lectura),
            "factura": info_factura if info_factura else None,
            "accion_factura": accion_factura,
            "message": f"Lectura actualizada. {mensaje_adicional or 'Sin factura asociada'}"
        }
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error al actualizar lectura: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al actualizar la lectura: {str(e)}"
        )


@router.delete("/{id_lectura}", status_code=status.HTTP_200_OK)
def eliminar_lectura(
    id_lectura: int,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Elimina la lectura si no tiene relaciones
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "eliminar")
    
    lectura = db.query(Lectura).filter(Lectura.id_lectura == id_lectura).first()
    
    if not lectura:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lectura no encontrada"
        )
    
    try:
        db.delete(lectura)
        db.commit()
        
        registrar_auditoria(
            db=db,
            accion="DELETE",
            descripcion=f"Lectura {id_lectura} eliminada por '{payload['sub']}'",
            id_usuario=current_user.id_usuario_sistema
        )
        
        registrar_notificacion(
            db=db,
            id_usuario=current_user.id_usuario_sistema,
            titulo="Lectura eliminada",
            mensaje=f"La lectura fue eliminada correctamente.",
            tipo="info"
        )
        
        return {
            "success": True,
            "accion": "eliminado",
            "message": "Lectura eliminada correctamente."
        }
    
    except IntegrityError:
        db.rollback()
        return {
            "success": False,
            "accion": "no_eliminado",
            "message": "⚠️ NO se puede eliminar la lectura porque está relacionada con otros módulos."
        }
    
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error inesperado al intentar eliminar la lectura."
        )


@router.patch("/{id_lectura}/toggle-status", response_model=dict)
def toggle_lectura_status(
    id_lectura: int,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Activa/Desactiva una lectura
    Requiere permiso: lecturas.actualizar o lecturas.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "actualizar")
    
    lectura = db.query(Lectura).filter(Lectura.id_lectura == id_lectura).first()
    
    if not lectura:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lectura no encontrada"
        )
    
    # Cambiar estado
    lectura.activo = not lectura.activo
    estado_texto = "activada" if lectura.activo else "desactivada"
    
    try:
        db.commit()
        db.refresh(lectura)
        
        # Auditoría
        registrar_auditoria(
            db=db,
            accion="UPDATE",
            descripcion=f"Lectura {id_lectura} fue {estado_texto} por '{payload['sub']}'",
            id_usuario=current_user.id_usuario_sistema
        )
        
        return lectura_to_response(lectura)
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error al cambiar estado: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al cambiar el estado de la lectura"
        )

# ========================================
# EXPORTAR PLANTILLA EXCEL OPTIMIZADA
# ========================================
# ========================================
# EXPORTAR PLANTILLA EXCEL OPTIMIZADA
# ========================================

@router.get("/export/template")
def exportar_plantilla(
    mes: Optional[int] = Query(None, ge=1, le=12, description="Mes del periodo (1-12)"),
    anio: Optional[int] = Query(None, ge=2020, description="Año del periodo"),
    incluir_todos: bool = Query(False, description="Incluir medidores con lectura registrada"),
    payload: dict = Depends(verify_token),
    db: Session = Depends(get_db)
):
    """
    Descarga plantilla Excel optimizada:
    - Excluye medidores con lectura en el periodo (por defecto)
    - Si todos tienen lectura, genera Excel informativo
    - incluir_todos=True: Permite descargar todos (para actualizar)
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "lectura")
    
    try:
        from sqlalchemy import func
        
        print(f"\n{'='*60}")
        print(f"📊 GENERANDO PLANTILLA OPTIMIZADA")
        print(f"{'='*60}")
        
        periodo_info = ""
        if mes and anio:
            periodo_info = f" {mes:02d}/{anio}"
            print(f"📅 Periodo: {periodo_info}")
            print(f"🔧 Incluir todos: {incluir_todos}")
        
        # ============================================
        # PASO 1: OBTENER ESTADÍSTICAS DEL PERIODO
        # ============================================
        medidores_con_lectura_ids = set()
        total_medidores_activos = 0
        
        if mes and anio:
            periodo = construir_periodo_consumo(anio, mes)
            # Medidores con lectura en el periodo
            lecturas_periodo = db.query(Lectura.id_medidor).filter(
                Lectura.periodo_consumo == periodo
            ).distinct().all()
            
            medidores_con_lectura_ids = {lectura[0] for lectura in lecturas_periodo}
            
            # Total de medidores activos
            total_medidores_activos = db.query(func.count(Medidor.id_medidor)).filter(
                Medidor.activo == True,
                Medidor.id_usuario_afi.isnot(None)
            ).scalar()
            
            print(f"📊 Medidores activos: {total_medidores_activos}")
            print(f"✅ Con lectura: {len(medidores_con_lectura_ids)}")
            print(f"⚠️ Sin lectura: {total_medidores_activos - len(medidores_con_lectura_ids)}")
        
        # ============================================
        # PASO 2: CONSULTA OPTIMIZADA (1 QUERY)
        # ============================================
        subquery_ultima_lectura = (
            db.query(
                Lectura.id_medidor,
                func.max(Lectura.id_lectura).label('max_id_lectura')
            )
            .group_by(Lectura.id_medidor)
            .subquery()
        )
        
        query = (
            db.query(
                Medidor.id_medidor,
                Medidor.num_medidor,
                UsuarioAfiliado.cod_usuario_afi,
                UsuarioSistema.nombres,
                UsuarioSistema.apellidos,
                Sector.nombre_sector,
                Lectura.lectura_actual.label('lectura_anterior')
            )
            .join(
                UsuarioAfiliado,
                UsuarioAfiliado.id_usuario_afi == Medidor.id_usuario_afi
            )
            .join(
                UsuarioSistema,
                UsuarioSistema.id_usuario_sistema == UsuarioAfiliado.id_usuario_sistema
            )
            .outerjoin(
                Sector,
                Sector.id_sector == Medidor.id_sector
            )
            .outerjoin(
                subquery_ultima_lectura,
                subquery_ultima_lectura.c.id_medidor == Medidor.id_medidor
            )
            .outerjoin(
                Lectura,
                Lectura.id_lectura == subquery_ultima_lectura.c.max_id_lectura
            )
            .filter(
                Medidor.activo == True,
                Medidor.id_usuario_afi.isnot(None)
            )
        )
        
        # ============================================
        # PASO 3: APLICAR FILTRO SEGÚN incluir_todos
        # ============================================
        if not incluir_todos and medidores_con_lectura_ids:
            query = query.filter(Medidor.id_medidor.notin_(medidores_con_lectura_ids))
        
        medidores_data = query.order_by(Medidor.num_medidor).all()
        
        print(f"✅ Medidores a exportar: {len(medidores_data)}")
        
        # ============================================
        # PASO 4: VALIDAR SI PERIODO ESTÁ COMPLETO
        # ============================================
        todos_tienen_lectura = (
            mes and anio and 
            len(medidores_con_lectura_ids) == total_medidores_activos and
            total_medidores_activos > 0 and
            not incluir_todos
        )
        
        if len(medidores_data) == 0 and todos_tienen_lectura:
            print("⚠️ PERIODO COMPLETO - Generando Excel informativo")
            return generar_excel_informativo(
                mes=mes,
                anio=anio,
                total_medidores=total_medidores_activos,
                total_con_lectura=len(medidores_con_lectura_ids),
                current_user=current_user,
                db=db
            )
        
        if len(medidores_data) == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No hay medidores activos disponibles"
            )
        
        # ============================================
        # PASO 5: GENERAR EXCEL NORMAL
        # ============================================
        wb = Workbook()
        ws_plantilla = wb.active
        ws_plantilla.title = "Plantilla Lecturas"
        
        # Encabezados
        headers = [
            "num_medidor", "sector", "codigo_afiliado",
            "nombre_afiliado", "lectura_anterior", "lectura_actual", "observacion"
        ]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_plantilla.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.protection = Protection(locked=True)
        
        # Anchos de columna
        column_widths = [20, 25, 18, 35, 18, 18, 40]
        for col, width in zip("ABCDEFG", column_widths):
            ws_plantilla.column_dimensions[col].width = width
        
        # Agregar datos
        for row_num, medidor_row in enumerate(medidores_data, 2):
            (id_medidor, num_medidor, cod_usuario_afi, nombres,
             apellidos, nombre_sector, lectura_anterior) = medidor_row
            
            codigo_afiliado = cod_usuario_afi or "N/A"
            nombre_afiliado = f"{nombres or ''} {apellidos or ''}".strip() or "Sin afiliado"
            sector_nombre = nombre_sector or "Sin sector"
            lectura_ant = lectura_anterior or 0
            
            # Columnas bloqueadas
            ws_plantilla.cell(row=row_num, column=1, value=num_medidor).protection = Protection(locked=True)
            ws_plantilla.cell(row=row_num, column=2, value=sector_nombre).protection = Protection(locked=True)
            ws_plantilla.cell(row=row_num, column=3, value=codigo_afiliado).protection = Protection(locked=True)
            ws_plantilla.cell(row=row_num, column=4, value=nombre_afiliado).protection = Protection(locked=True)
            ws_plantilla.cell(row=row_num, column=5, value=lectura_ant).protection = Protection(locked=True)
            
            # Columnas editables
            ws_plantilla.cell(row=row_num, column=6, value="").protection = Protection(locked=False)
            ws_plantilla.cell(row=row_num, column=7, value="").protection = Protection(locked=False)
        
        ws_plantilla.protection.sheet = True
        ws_plantilla.protection.enable()
        
        # ============================================
        # HOJA DE INSTRUCCIONES
        # ============================================
        ws_instrucciones = wb.create_sheet("Instrucciones")
        periodo_texto = f"{mes:02d}/{anio}" if mes and anio else "N/A"
        porcentaje = (len(medidores_con_lectura_ids) / total_medidores_activos * 100) if total_medidores_activos > 0 else 0
        
        instrucciones = [
            ["📋 INSTRUCCIONES PARA CARGA MASIVA DE LECTURAS"],
            [""],
            [f"📅 PERIODO: {periodo_texto}"],
            [f"📊 PROGRESO DEL PERIODO: {porcentaje:.1f}% completado"],
            [f"✅ Lecturas registradas: {len(medidores_con_lectura_ids)} de {total_medidores_activos}"],
            [f"📝 Medidores en esta plantilla: {len(medidores_data)}"],
            [""],
            ["1️⃣ USO DE LA PLANTILLA:"],
            [" • Complete SOLO 'lectura_actual' y 'observacion'"],
            [" • Las demás columnas están bloqueadas"],
            [" • lectura_actual debe ser >= lectura_anterior"],
            [""],
            ["2️⃣ PROCESO:"],
            [" • Guarde el archivo después de completar"],
            [" • Suba usando 'Crear desde Excel'"],
            [" • El sistema validará los datos"],
            [""],
            [f"📅 Generada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
            [f"👤 Usuario: {current_user.nombres} {current_user.apellidos}"],
        ]
        
        for row_num, fila in enumerate(instrucciones, 1):
            cell = ws_instrucciones.cell(row=row_num, column=1, value=fila[0])
            if "📋" in str(fila[0]):
                cell.font = Font(size=14, bold=True, color="4472C4")
            elif any(emoji in str(fila[0]) for emoji in ["1️⃣", "2️⃣", "📅", "📊"]):
                cell.font = Font(size=12, bold=True)
        
        ws_instrucciones.column_dimensions['A'].width = 80
        
        # ============================================
        # GUARDAR Y RETORNAR
        # ============================================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = output.getvalue()
        output.close()
        
        print(f"✅ Excel generado: {len(excel_data)} bytes\n")
        
        registrar_auditoria(
            db=db,
            accion="DOWNLOAD_TEMPLATE",
            descripcion=f"Plantilla{periodo_info} - {len(medidores_data)} medidores ({porcentaje:.1f}% completo)",
            id_usuario=current_user.id_usuario_sistema
        )
        
        filename = f"plantilla_lecturas_{anio}{mes:02d}_{datetime.now().strftime('%H%M%S')}.xlsx" if mes and anio else f"plantilla_lecturas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(len(excel_data))
            }
        )
    
    except HTTPException as he:
        raise he
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar plantilla: {str(e)}"
        )


# ============================================
# FUNCIÓN: GENERAR EXCEL INFORMATIVO
# ============================================
def generar_excel_informativo(
    mes: int,
    anio: int,
    total_medidores: int,
    total_con_lectura: int,
    current_user,
    db: Session
):
    """
    Genera Excel informativo cuando el periodo está 100% completo
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import io
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Periodo Completo"
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # ============================================
    # TÍTULO PRINCIPAL
    # ============================================
    ws.merge_cells('A1:F1')
    cell_titulo = ws['A1']
    cell_titulo.value = "✅ PERIODO COMPLETO"
    cell_titulo.font = Font(size=18, bold=True, color="FFFFFF")
    cell_titulo.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # ============================================
    # INFORMACIÓN DEL PERIODO
    # ============================================
    ws.merge_cells('A3:F3')
    cell_periodo = ws['A3']
    cell_periodo.value = f"📅 Periodo: {mes:02d}/{anio}"
    cell_periodo.font = Font(size=14, bold=True, color="1F2937")
    cell_periodo.alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 25
    
    # ============================================
    # ESTADÍSTICAS CON DISEÑO
    # ============================================
    row = 5
    ws.merge_cells(f'A{row}:F{row}')
    cell_stats_title = ws[f'A{row}']
    cell_stats_title.value = "📊 ESTADÍSTICAS DEL PERIODO"
    cell_stats_title.font = Font(size=13, bold=True, color="FFFFFF")
    cell_stats_title.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    cell_stats_title.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 25
    
    row += 2
    stats_data = [
        ("Total de medidores activos", total_medidores, "4B5563"),
        ("Lecturas registradas", total_con_lectura, "059669"),
        ("Medidores pendientes", 0, "6B7280"),
        ("Porcentaje completado", "100%", "059669"),
    ]
    
    for label, value, color in stats_data:
        ws.merge_cells(f'B{row}:D{row}')
        cell_label = ws[f'B{row}']
        cell_label.value = label
        cell_label.font = Font(size=11, bold=True)
        cell_label.alignment = Alignment(horizontal="left")
        cell_label.border = thin_border
        
        ws.merge_cells(f'E{row}:F{row}')
        cell_value = ws[f'E{row}']
        cell_value.value = value
        cell_value.font = Font(size=11, bold=True, color=color)
        cell_value.alignment = Alignment(horizontal="right")
        cell_value.border = thin_border
        
        ws.row_dimensions[row].height = 22
        row += 1
    
    # ============================================
    # BARRA DE PROGRESO VISUAL
    # ============================================
    row += 1
    ws.merge_cells(f'B{row}:F{row}')
    cell_progress = ws[f'B{row}']
    cell_progress.value = "█" * 30 + " 100%"
    cell_progress.font = Font(size=10, color="059669")
    cell_progress.alignment = Alignment(horizontal="center")
    
    # ============================================
    # OPCIONES DISPONIBLES
    # ============================================
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell_opciones = ws[f'A{row}']
    cell_opciones.value = "💡 ¿QUÉ PUEDES HACER AHORA?"
    cell_opciones.font = Font(size=13, bold=True, color="FFFFFF")
    cell_opciones.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    cell_opciones.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 25
    
    row += 2
    opciones = [
        ("1️⃣", "Ver lecturas registradas", "Consulta el módulo de lecturas con el filtro de este periodo"),
        ("2️⃣", "Generar reportes", "Exporta reportes de consumo y facturación del periodo"),
        ("3️⃣", "Generar facturas", "Procede a facturar las lecturas del periodo"),
        ("4️⃣", "Seleccionar otro periodo", "Trabaja con un periodo diferente que aún tenga pendientes"),
    ]
    
    for emoji, titulo, descripcion in opciones:
        # Emoji
        cell_emoji = ws[f'B{row}']
        cell_emoji.value = emoji
        cell_emoji.font = Font(size=14)
        cell_emoji.alignment = Alignment(horizontal="center", vertical="center")
        
        # Título
        ws.merge_cells(f'C{row}:D{row}')
        cell_titulo = ws[f'C{row}']
        cell_titulo.value = titulo
        cell_titulo.font = Font(bold=True, size=11, color="1F2937")
        cell_titulo.alignment = Alignment(horizontal="left", vertical="center")
        
        # Descripción
        ws.merge_cells(f'E{row}:F{row}')
        cell_desc = ws[f'E{row}']
        cell_desc.value = descripcion
        cell_desc.font = Font(size=9, color="6B7280")
        cell_desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        ws.row_dimensions[row].height = 30
        row += 2
    
    # ============================================
    # NOTA INFORMATIVA
    # ============================================
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell_nota = ws[f'A{row}']
    cell_nota.value = "ℹ️ Este periodo ya no requiere carga de lecturas. Todas las lecturas están completas."
    cell_nota.font = Font(size=10, italic=True, color="6B7280")
    cell_nota.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    cell_nota.alignment = Alignment(horizontal="center", vertical="center")
    cell_nota.border = thin_border
    ws.row_dimensions[row].height = 25
    
    # ============================================
    # FOOTER
    # ============================================
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell_footer = ws[f'A{row}']
    cell_footer.value = f"📅 Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | 👤 Usuario: {current_user.nombres} {current_user.apellidos}"
    cell_footer.font = Font(size=9, color="9CA3AF")
    cell_footer.alignment = Alignment(horizontal="center")
    
    # ============================================
    # AJUSTAR ANCHOS DE COLUMNA
    # ============================================
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    
    # ============================================
    # GUARDAR
    # ============================================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    excel_data = output.getvalue()
    output.close()
    
    print(f"✅ Excel informativo generado: {len(excel_data)} bytes\n")
    
    registrar_auditoria(
        db=db,
        accion="DOWNLOAD_TEMPLATE_COMPLETE",
        descripcion=f"Excel informativo {mes:02d}/{anio} - Periodo 100% completo",
        id_usuario=current_user.id_usuario_sistema
    )
    
    filename = f"periodo_completo_{anio}{mes:02d}_{datetime.now().strftime('%H%M%S')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(excel_data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Length": str(len(excel_data))
        }
    )


# =================================================
# IMPORTAR LECTURAS DESDE EXCEL - CREAR DESDE EXCEL 
# =================================================

@router.post("/import/excel", response_model=LecturaBulkResponse, status_code=status.HTTP_201_CREATED)
async def importar_lecturas_excel(
    file: UploadFile = File(...),
    payload: dict = Depends(verify_token),
    db: Session = Depends(get_db)
):
    """
    Importa lecturas desde un archivo Excel
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "crear")
    
    exitosos = []
    fallidos = []
    
    try:
        # Leer el archivo Excel
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        print(f"\n{'='*60}")
        print(f"🚀 INICIANDO IMPORTACIÓN DE LECTURAS DESDE EXCEL")
        print(f"{'='*60}\n")
        
        fecha_lectura = date.today()
        periodo_consumo = periodo_desde_fecha(fecha_lectura)
        
        # Procesar cada fila (empezar desde la 2 para saltar encabezados)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Extraer valores (nuevas columnas)
                num_medidor = row[0]
                sector = row[1]
                codigo_UsuarioAfiliado = row[2]
                nombre_UsuarioAfiliado = row[3]
                lectura_anterior = row[4]
                lectura_actual = row[5]
                observacion = row[6] if len(row) > 6 else None
                
                # Validar datos
                if not num_medidor or not lectura_actual:
                    continue  # Saltar filas vacías
                
                lectura_actual = int(lectura_actual)
                lectura_anterior = int(lectura_anterior) if lectura_anterior else 0
                
                # Validar que la lectura actual sea mayor o igual a la anterior
                if lectura_actual < lectura_anterior:
                    raise ValueError(f"La lectura actual ({lectura_actual}) es menor que la anterior ({lectura_anterior})")
                
                # Calcular consumo
                consumo_m3 = lectura_actual - lectura_anterior
                
                # Buscar el medidor por número
                medidor = db.query(Medidor).filter(
                    Medidor.num_medidor == str(num_medidor).strip()
                ).first()
                
                if not medidor:
                    raise ValueError(f"Medidor '{num_medidor}' no encontrado en el sistema")

                lectura_existente = db.query(Lectura).filter(
                    Lectura.id_medidor == medidor.id_medidor,
                    Lectura.periodo_consumo == periodo_consumo,
                    Lectura.activo == True
                ).first()

                if lectura_existente:
                    raise ValueError(f"Ya existe lectura para el periodo de consumo {periodo_consumo}")
                
                # Crear lectura
                nueva_lectura = Lectura(
                    id_medidor=medidor.id_medidor,
                    lectura_actual=lectura_actual,
                    lectura_anterior=lectura_anterior,
                    consumo_m3=consumo_m3,
                    fecha_lectura=fecha_lectura,
                    periodo_consumo=periodo_consumo,
                    id_lector=current_user.id_usuario_sistema,
                    observacion=observacion.strip() if observacion else None,
                    activo=True
                )
                
                db.add(nueva_lectura)
                db.flush()
                
                exitosos.append(LecturaBulkResult(
                    fila=row_num,
                    id_medidor=medidor.id_medidor,
                    num_medidor=medidor.num_medidor,
                    lectura_anterior=lectura_anterior,
                    lectura_actual=lectura_actual,
                    consumo_m3=consumo_m3,
                    id_lectura=nueva_lectura.id_lectura
                ))
                
                print(f"✅ Fila {row_num}: Lectura creada para medidor {medidor.num_medidor} - Consumo: {consumo_m3}m³")
                
            except Exception as e:
                fallidos.append(LecturaBulkError(
                    fila=row_num,
                    id_medidor=None,
                    num_medidor=num_medidor if 'num_medidor' in locals() else None,
                    error=str(e)
                ))
                print(f"❌ Fila {row_num}: Error - {str(e)}")
        
        # Commit si hubo éxitos
        if exitosos:
            db.commit()
            
            # Auditoría
            registrar_auditoria(
                db=db,
                accion="IMPORT_EXCEL",
                descripcion=f"Importación masiva de lecturas: {len(exitosos)} exitosos, {len(fallidos)} fallidos por '{current_user.usuario}'",
                id_usuario=current_user.id_usuario_sistema
            )
            
            # Notificación
            registrar_notificacion(
                db=db,
                id_usuario=current_user.id_usuario_sistema,
                titulo="Importación de lecturas completada",
                mensaje=f"Se importaron {len(exitosos)} lecturas correctamente. Errores: {len(fallidos)}",
                tipo="exito"
            )
        
        print(f"\n{'='*60}")
        print(f"✅ IMPORTACIÓN COMPLETADA")
        print(f"Total procesados: {len(exitosos) + len(fallidos)}")
        print(f"Exitosos: {len(exitosos)}")
        print(f"Fallidos: {len(fallidos)}")
        print(f"{'='*60}\n")
        
        return LecturaBulkResponse(
            exitosos=exitosos,
            fallidos=fallidos,
            total_procesados=len(exitosos) + len(fallidos),
            total_exitosos=len(exitosos),
            total_fallidos=len(fallidos)
        )
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error en importación: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al importar lecturas: {str(e)}"
        )


# ========================================
# EXPORTAR LECTURAS A EXCEL
# ========================================

@router.get("/export/excel")
def exportar_lecturas_excel(
    mes: Optional[int] = Query(None, ge=1, le=12),
    anio: Optional[int] = Query(None, ge=2020),
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    payload: dict = Depends(verify_token),
    db: Session = Depends(get_db)
):
    """
    Exporta lecturas a Excel con filtros opcionales
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "lectura")
    
    try:
        # Consultar lecturas
        query = db.query(Lectura).filter(Lectura.activo == True)
        
        if mes is not None and anio is not None:
            query = query.filter(Lectura.periodo_consumo == construir_periodo_consumo(anio, mes))
        if fecha_desde:
            query = query.filter(Lectura.fecha_lectura >= fecha_desde)
        if fecha_hasta:
            query = query.filter(Lectura.fecha_lectura <= fecha_hasta)
        
        lecturas = query.order_by(
            Lectura.periodo_consumo.desc(),
            Lectura.fecha_lectura.desc()
        ).all()
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Lecturas"
        
        # Encabezados
        headers = [
            "ID Lectura",
            "Medidor",
            "Lectura Anterior",
            "Lectura Actual",
            "Consumo (m³)",
            "Periodo Consumo",
            "Fecha Lectura",
            "Lector",
            "Observación"
        ]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Datos
        for row_num, lectura in enumerate(lecturas, 2):
            # Obtener información del medidor
            medidor = lectura.medidor
            num_medidor = medidor.num_medidor if medidor else "N/A"
            
            # Obtener información del lector
            lector = lectura.lector
            nombre_lector = f"{lector.nombres} {lector.apellidos}" if lector else "N/A"
            
            ws.cell(row=row_num, column=1, value=lectura.id_lectura)
            ws.cell(row=row_num, column=2, value=num_medidor)
            ws.cell(row=row_num, column=3, value=lectura.lectura_anterior)
            ws.cell(row=row_num, column=4, value=lectura.lectura_actual)
            ws.cell(row=row_num, column=5, value=lectura.consumo_m3)
            ws.cell(row=row_num, column=6, value=lectura.periodo_consumo)
            ws.cell(row=row_num, column=7, value=lectura.fecha_lectura.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=8, value=nombre_lector)
            ws.cell(row=row_num, column=9, value=lectura.observacion or "")
        
        # Ajustar anchos
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            ws.column_dimensions[col].width = 20
        
        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_data = output.getvalue()
        output.close()
        
        # Registrar auditoría
        registrar_auditoria(
            db=db,
            accion="EXPORT_EXCEL",
            descripcion=f"Lecturas exportadas a Excel por '{current_user.usuario}' ({len(lecturas)} registros)",
            id_usuario=current_user.id_usuario_sistema
        )
        
        filename = f"lecturas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        final_output = io.BytesIO(excel_data)
        
        return StreamingResponse(
            final_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(len(excel_data))
            }
        )
    
    except Exception as e:
        print(f"❌ Error exportando lecturas: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al exportar lecturas: {str(e)}"
        )
    
#
# Intentar configurar locale español
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')
    except:
        pass  # Usar nombres de meses en inglés como fallback

# Diccionario de nombres de meses en español (fallback)
MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}


# ========================================
# 🆕 ENDPOINT: OBTENER PERIODOS DISPONIBLES
# ========================================
@router.get("/periodos/disponibles", response_model=dict)
def obtener_periodos_disponibles(
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Obtiene los periodos (mes/año) disponibles para cargar lecturas.
    Muestra:
    - Periodo actual sugerido
    - Últimos 6 meses con estadísticas
    - Próximos 2 meses
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "lectura")
    
    try:
        # Fecha actual
        hoy = date.today()
        mes_actual = hoy.month
        anio_actual = hoy.year
        
        # ✅ Total de medidores activos CON AFILIADOS
        total_medidores = db.query(func.count(Medidor.id_medidor)).filter(
            Medidor.activo == True,
            Medidor.id_usuario_afi.isnot(None)  # Solo medidores con afiliado
        ).scalar() or 0
        
        periodos = []
        
        # Generar últimos 6 meses + mes actual + próximos 2 meses
        for offset in range(-6, 3):
            # Calcular mes y año
            fecha_temp = date(anio_actual, mes_actual, 1)
            
            # Sumar/restar meses
            mes_temp = mes_actual + offset
            anio_temp = anio_actual
            
            while mes_temp > 12:
                mes_temp -= 12
                anio_temp += 1
            while mes_temp < 1:
                mes_temp += 12
                anio_temp -= 1
            
            # Contar lecturas del periodo
            periodo_temp = construir_periodo_consumo(anio_temp, mes_temp)
            total_lecturas_periodo = db.query(func.count(Lectura.id_lectura)).filter(
                Lectura.periodo_consumo == periodo_temp,
                Lectura.activo == True
            ).scalar() or 0
            
            # Determinar si es sugerido (mes actual o siguiente si ya tiene muchas lecturas)
            porcentaje = (total_lecturas_periodo / total_medidores * 100) if total_medidores > 0 else 0
            sugerido = False
            
            if mes_temp == mes_actual and anio_temp == anio_actual:
                sugerido = True  # Mes actual siempre sugerido
            elif offset == 1 and porcentaje < 80:  # Mes siguiente si el actual está completo
                sugerido = True
            
            periodos.append({
                "mes": mes_temp,
                "anio": anio_temp,
                "nombre_mes": MESES_ES.get(mes_temp, f"Mes {mes_temp}"),
                "tiene_lecturas": total_lecturas_periodo > 0,
                "total_lecturas": total_lecturas_periodo,
                "total_medidores": total_medidores,
                "porcentaje_completado": round(porcentaje, 1),
                "sugerido": sugerido
            })
        
        # Ordenar por año y mes descendente (más reciente primero)
        periodos.sort(key=lambda x: (x["anio"], x["mes"]), reverse=True)
        
        # Identificar periodo actual
        periodo_actual = next((p for p in periodos if p["sugerido"]), periodos[0])
        
        return {
            "periodo_actual": periodo_actual,
            "periodos_disponibles": periodos,
            "total_medidores_activos": total_medidores
        }
    
    except Exception as e:
        print(f"❌ Error obteniendo periodos: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener periodos disponibles: {str(e)}"
        )


# ========================================
# 🆕 IMPORTAR LECTURAS CON PERIODO - CREAR DESDE EXCEL
# ========================================
@router.post("/import/excel/periodo", response_model=LecturaBulkResponse, status_code=status.HTTP_201_CREATED)
async def importar_lecturas_excel_con_periodo(
    file: UploadFile = File(...),
    mes: int = Form(..., ge=1, le=12, description="Mes de las lecturas"),
    anio: int = Form(..., ge=2020, description="Año de las lecturas"),
    payload: dict = Depends(verify_token),
    db: Session = Depends(get_db)
):
    """
    Importa lecturas desde Excel con periodo específico.
    La tarifa se determina automáticamente según el consumo.
    Genera facturas automáticamente para cada lectura exitosa.
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "crear")
    
    exitosos = []
    fallidos = []
    facturas_generadas = 0
    
    try:
        # Crear fecha del periodo (primer día del mes)
        fecha_lectura = date.today()
        periodo_consumo = construir_periodo_consumo(anio, mes)
        
        print(f"\n{'='*60}")
        print(f"🚀 IMPORTACIÓN PARA PERIODO: {MESES_ES.get(mes, mes)}/{anio}")
        print(f"{'='*60}\n")
        
        # Leer Excel
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Procesar filas
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                num_medidor = row[0]
                lectura_anterior = row[4]
                lectura_actual = row[5]
                observacion = row[6] if len(row) > 6 else None
                
                if not num_medidor or not lectura_actual:
                    continue
                
                lectura_actual = int(lectura_actual)
                lectura_anterior = int(lectura_anterior) if lectura_anterior else 0
                
                if lectura_actual < lectura_anterior:
                    raise ValueError(f"Lectura actual ({lectura_actual}) menor que anterior ({lectura_anterior})")
                
                consumo_m3 = lectura_actual - lectura_anterior
                
                # Buscar medidor
                medidor = db.query(Medidor).filter(
                    Medidor.num_medidor == str(num_medidor).strip()
                ).first()
                
                if not medidor:
                    raise ValueError(f"Medidor '{num_medidor}' no encontrado")
                
                # Validar duplicado
                lectura_existente = db.query(Lectura).filter(
                    Lectura.id_medidor == medidor.id_medidor,
                    Lectura.periodo_consumo == periodo_consumo,
                    Lectura.activo == True
                ).first()
                
                if lectura_existente:
                    raise ValueError(f"Ya existe lectura para {MESES_ES.get(mes, mes)}/{anio}")
                
                # Crear lectura
                nueva_lectura = Lectura(
                    id_medidor=medidor.id_medidor,
                    lectura_actual=lectura_actual,
                    lectura_anterior=lectura_anterior,
                    consumo_m3=consumo_m3,
                    fecha_lectura=fecha_lectura,
                    periodo_consumo=periodo_consumo,
                    id_lector=current_user.id_usuario_sistema,
                    observacion=observacion.strip() if observacion else None,
                    activo=True
                )
                
                db.add(nueva_lectura)
                db.flush()
                
                # Generar factura
                exito_factura, mensaje_factura, factura = generar_factura_desde_lectura(
                    db=db,
                    lectura=nueva_lectura,
                    aplicar_servicios=True,
                    aplicar_multas=True
                )
                
                if exito_factura:
                    facturas_generadas += 1
                    print(f"✅ Fila {row_num}: {medidor.num_medidor} - {consumo_m3}m³ | Factura: {factura.num_factura}")
                else:
                    print(f"⚠️  Fila {row_num}: Lectura OK pero factura falló: {mensaje_factura}")
                
                exitosos.append(LecturaBulkResult(
                    fila=row_num,
                    id_medidor=medidor.id_medidor,
                    num_medidor=medidor.num_medidor,
                    lectura_anterior=lectura_anterior,
                    lectura_actual=lectura_actual,
                    consumo_m3=consumo_m3,
                    id_lectura=nueva_lectura.id_lectura
                ))
                
            except Exception as e:
                fallidos.append(LecturaBulkError(
                    fila=row_num,
                    id_medidor=None,
                    num_medidor=num_medidor if 'num_medidor' in locals() else None,
                    error=str(e)
                ))
                print(f"❌ Fila {row_num}: {str(e)}")
        
        # Commit
        if exitosos:
            db.commit()
            
            registrar_auditoria(
                db=db,
                accion="IMPORT_EXCEL",
                descripcion=f"Importación {MESES_ES.get(mes, mes)}/{anio}: {len(exitosos)} lecturas, {facturas_generadas} facturas generadas",
                id_usuario=current_user.id_usuario_sistema
            )
            
            registrar_notificacion(
                db=db,
                id_usuario=current_user.id_usuario_sistema,
                titulo=f"Lecturas {MESES_ES.get(mes, mes)}/{anio} importadas",
                mensaje=f"{len(exitosos)} lecturas y {facturas_generadas} facturas generadas correctamente",
                tipo="exito"
            )
        
        print(f"\n{'='*60}")
        print(f"✅ COMPLETADO")
        print(f"   Lecturas exitosas: {len(exitosos)}")
        print(f"   Facturas generadas: {facturas_generadas}")
        print(f"   Fallidos: {len(fallidos)}")
        print(f"{'='*60}\n")
        
        return LecturaBulkResponse(
            exitosos=exitosos,
            fallidos=fallidos,
            total_procesados=len(exitosos) + len(fallidos),
            total_exitosos=len(exitosos),
            total_fallidos=len(fallidos)
        )
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al importar: {str(e)}"
        )


# ========================================
# ENDPOINT: GENERAR LECTURAS ESTIMADA
# ========================================

@router.post("/generar-estimadas", response_model=dict)
def generar_lecturas_estimadas(
    mes: int = Query(..., ge=1, le=12, description="Mes para generar lecturas"),
    anio: int = Query(..., ge=2020, description="Año para generar lecturas"),
    meses_promedio: int = Query(3, ge=1, le=12, description="Meses para calcular promedio"),
    consumo_default: int = Query(10, ge=0, description="Consumo por defecto para medidores sin historial (m³)"),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Genera lecturas estimadas para medidores que NO tienen lectura en el período especificado.
    
    CASOS MANEJADOS:
    1. Medidor CON historial: Calcula promedio de últimos N meses
    2. Medidor SIN historial: Usa lectura_anterior del medidor + consumo_default
    3. Medidor nuevo (lectura_anterior = 0): Genera lectura inicial con consumo_default
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "crear")
    
    try:
        periodo_consumo = construir_periodo_consumo(anio, mes)
        # 1. Obtener todos los medidores activos con usuarios
        medidores_con_usuario = db.query(Medidor).filter(
            Medidor.activo == True,
            Medidor.id_usuario_afi.isnot(None)
        ).all()
        
        if not medidores_con_usuario:
            return {
                "success": False,
                "message": "No hay medidores con usuarios asignados"
            }
        
        # 2. Obtener medidores que YA tienen lectura en el período
        medidores_con_lectura = db.query(Lectura.id_medidor).filter(
            Lectura.periodo_consumo == periodo_consumo,
            Lectura.activo == True
        ).distinct().all()
        
        ids_con_lectura = {m[0] for m in medidores_con_lectura}
        
        # 3. Filtrar medidores SIN lectura en el período
        medidores_sin_lectura = [
            m for m in medidores_con_usuario 
            if m.id_medidor not in ids_con_lectura
        ]
        
        if not medidores_sin_lectura:
            return {
                "success": True,
                "message": "Todos los medidores ya tienen lectura registrada",
                "lecturas_generadas": 0,
                "detalles": []
            }
        
        # 4. Calcular consumo promedio del sistema (para referencia)
        consumo_promedio_sistema = db.query(
            func.avg(Lectura.consumo_m3)
        ).filter(
            Lectura.activo == True,
            Lectura.es_estimada == False,
            Lectura.consumo_m3 > 0
        ).scalar() or consumo_default
        
        consumo_promedio_sistema = round(consumo_promedio_sistema)

        ids_sin_lectura = [m.id_medidor for m in medidores_sin_lectura]

        historial_rows = db.query(Lectura).filter(
            Lectura.id_medidor.in_(ids_sin_lectura),
            Lectura.activo == True,
            Lectura.es_estimada == False
        ).order_by(
            Lectura.id_medidor.asc(),
            Lectura.periodo_consumo.desc(),
            Lectura.fecha_lectura.desc(),
            Lectura.id_lectura.desc()
        ).all()

        historial_por_medidor = defaultdict(list)
        for lectura_hist in historial_rows:
            lecturas_medidor = historial_por_medidor[lectura_hist.id_medidor]
            if len(lecturas_medidor) < meses_promedio:
                lecturas_medidor.append(lectura_hist)

        ultima_rows = db.query(Lectura).filter(
            Lectura.id_medidor.in_(ids_sin_lectura),
            Lectura.activo == True
        ).order_by(
            Lectura.id_medidor.asc(),
            Lectura.periodo_consumo.desc(),
            Lectura.fecha_lectura.desc(),
            Lectura.id_lectura.desc()
        ).all()

        ultima_por_medidor = {}
        for lectura_hist in ultima_rows:
            ultima_por_medidor.setdefault(lectura_hist.id_medidor, lectura_hist)
        
        # 5. Generar lecturas estimadas
        lecturas_generadas = []
        lecturas_fallidas = []
        lecturas_pendientes = []
        
        for medidor in medidores_sin_lectura:
            try:
                # 🔹 CASO 1: Buscar historial de lecturas del medidor
                ultimas_lecturas = historial_por_medidor.get(medidor.id_medidor, [])
                
                # Variables para la lectura estimada
                lectura_anterior = 0
                consumo_estimado = 0
                metodo_calculo = ""
                
                if ultimas_lecturas:
                    # ✅ MEDIDOR CON HISTORIAL: Calcular promedio
                    consumo_estimado = sum(l.consumo_m3 for l in ultimas_lecturas) / len(ultimas_lecturas)
                    consumo_estimado = round(consumo_estimado)
                    lectura_anterior = ultimas_lecturas[0].lectura_actual
                    metodo_calculo = "historial de consumo"
                    
                else:
                    # 🔹 CASO 2: MEDIDOR SIN HISTORIAL
                    # Obtener última lectura conocida del endpoint /medidores/lista/completa
                    ultima_lectura_conocida = ultima_por_medidor.get(medidor.id_medidor)
                    
                    if ultima_lectura_conocida:
                        # ✅ Tiene una lectura previa (aunque sea antigua)
                        lectura_anterior = ultima_lectura_conocida.lectura_actual
                        consumo_estimado = consumo_default
                        metodo_calculo = "consumo sugerido sin historial reciente"
                    else:
                        # ✅ MEDIDOR COMPLETAMENTE NUEVO (primera lectura)
                        lectura_anterior = 0
                        consumo_estimado = consumo_default
                        metodo_calculo = "consumo inicial sugerido"
                
                # Calcular lectura estimada
                lectura_estimada = lectura_anterior + consumo_estimado
                
                # Crear lectura estimada
                nueva_lectura = Lectura(
                    id_medidor=medidor.id_medidor,
                    lectura_actual=lectura_estimada,
                    lectura_anterior=lectura_anterior,
                    consumo_m3=consumo_estimado,
                    fecha_lectura=date.today(),
                    periodo_consumo=periodo_consumo,
                    id_lector=current_user.id_usuario_sistema,
                    observacion=f"Lectura estimada por {metodo_calculo.lower()}",
                    activo=True,
                    es_estimada=True
                )
                
                
                # Obtener información del afiliado
                afiliado = medidor.usuario_afiliado
                nombre_afiliado = "Sin afiliado"
                codigo_afiliado = "N/A"
                if afiliado:
                    codigo_afiliado = afiliado.cod_usuario_afi or "N/A"
                    if afiliado.usuario_sistema:
                        us = afiliado.usuario_sistema
                        nombre_afiliado = f"{us.nombres} {us.apellidos}"
                
                lecturas_generadas.append({
                    "id_lectura": None,
                    "medidor": medidor.num_medidor,
                    "codigo_afiliado": codigo_afiliado,
                    "nombre_afiliado": nombre_afiliado,
                    "lectura_anterior": lectura_anterior,
                    "lectura_estimada": lectura_estimada,
                    "consumo_estimado": consumo_estimado,
                    "metodo_calculo": metodo_calculo,
                    "tiene_historial": len(ultimas_lecturas) > 0
                })
                lecturas_pendientes.append(nueva_lectura)
                
            except Exception as e:
                lecturas_fallidas.append({
                    "medidor": medidor.num_medidor,
                    "razon": str(e)
                })
                continue
        
        # 6. Confirmar transacción
        if lecturas_pendientes:
            db.add_all(lecturas_pendientes)
            db.flush()
            for detalle, lectura_creada in zip(lecturas_generadas, lecturas_pendientes):
                detalle["id_lectura"] = lectura_creada.id_lectura

        if lecturas_generadas:
            db.commit()
            
            # Contadores por tipo
            con_historial = sum(1 for l in lecturas_generadas if l["tiene_historial"])
            sin_historial = len(lecturas_generadas) - con_historial
            
            # Auditoría
            registrar_auditoria(
                db=db,
                accion="GENERAR_ESTIMADAS",
                descripcion=f"Generadas {len(lecturas_generadas)} lecturas estimadas para {MESES_ES.get(mes)}/{anio} - Con historial: {con_historial}, Sin historial: {sin_historial}",
                id_usuario=current_user.id_usuario_sistema
            )
        
        return {
            "success": True,
            "message": f"Proceso completado. Generadas {len(lecturas_generadas)} lecturas estimadas",
            "lecturas_generadas": len(lecturas_generadas),
            "lecturas_fallidas": len(lecturas_fallidas),
            "con_historial": sum(1 for l in lecturas_generadas if l["tiene_historial"]),
            "sin_historial": sum(1 for l in lecturas_generadas if not l["tiene_historial"]),
            "periodo": f"{MESES_ES.get(mes)} {anio}",
            "consumo_promedio_sistema": consumo_promedio_sistema,
            "detalles": lecturas_generadas,
            "fallidas": lecturas_fallidas
        }
        
    except Exception as e:
        db.rollback()
        print(f"❌ Error generando lecturas estimadas: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar lecturas estimadas: {str(e)}"
        )


# ========================================
# ENDPOINT PARA CONFIRMAR LECTURA ESTIMADA
# ========================================

@router.patch("/{id_lectura}/confirmar-estimada", response_model=dict)
def confirmar_lectura_estimada(
    id_lectura: int,
    lectura_real: int = Query(..., description="Lectura real tomada"),
    observacion: Optional[str] = Query(None, description="Observación adicional"),
    generar_factura: bool = Query(True, description="Generar factura automáticamente"),
    tipo_descuento: str = Query('ninguno', description="Tipo: ninguno/porcentaje/valor"),
    valor_descuento: float = Query(0.0, ge=0, description="Valor del descuento"),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Convierte una lectura estimada en lectura real con el valor correcto.
    Opcionalmente genera la factura automáticamente.
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "actualizar")
    
    try:
        lectura = db.query(Lectura).filter(
            Lectura.id_lectura == id_lectura
        ).first()
        
        if not lectura:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Lectura no encontrada"
            )
        
        if not lectura.es_estimada:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Esta lectura no es estimada"
            )
        
        # Validar que la lectura real sea mayor o igual a la anterior
        if lectura_real < lectura.lectura_anterior:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La lectura real no puede ser menor que la lectura anterior"
            )
        
        # Obtener información del medidor y afiliado
        medidor = lectura.medidor
        afiliado = medidor.usuario_afiliado if medidor else None
        usuario_afiliado = afiliado.usuario_sistema if afiliado else None
        
        if usuario_afiliado:
            nombre_afiliado = f"{usuario_afiliado.nombres} {usuario_afiliado.apellidos}"
            id_usuario_afiliado = usuario_afiliado.id_usuario_sistema
        else:
            nombre_afiliado = "Usuario desconocido"
            id_usuario_afiliado = None
        
        # Actualizar lectura
        lectura.lectura_actual = lectura_real
        lectura.consumo_m3 = lectura_real - lectura.lectura_anterior
        lectura.es_estimada = False  # Ya no es estimada
        lectura.id_lector = current_user.id_usuario_sistema
        
        if observacion:
            lectura.observacion = f"Lectura confirmada desde estimación. {observacion.strip()}"
        else:
            lectura.observacion = "Lectura confirmada desde estimación"
        
        db.flush()  # Guardar cambios de lectura antes de generar factura
        
        # ✅ GENERAR FACTURA AUTOMÁTICAMENTE
        factura_generada = None
        mensaje_factura = ""
        
        if generar_factura:
            exito, mensaje, factura_generada = generar_factura_desde_lectura(
                db=db,
                lectura=lectura,
                tipo_descuento=tipo_descuento,
                valor_descuento=valor_descuento,
                aplicar_servicios=True,
                aplicar_multas=True
            )
            
            if exito:
                mensaje_factura = "Factura generada correctamente."
            else:
                mensaje_factura = f"No se pudo generar la factura: {mensaje}"
        
        # Commit final
        db.commit()
        db.refresh(lectura)
        
        # Auditoría
        registrar_auditoria(
            db=db,
            accion="CONFIRMAR_ESTIMADA",
            descripcion=f"Lectura {id_lectura} confirmada - Real: {lectura_real}m³ - {mensaje_factura}",
            id_usuario=current_user.id_usuario_sistema
        )
        
        # Notificación para el lector
        registrar_notificacion(
            db=db,
            id_usuario=current_user.id_usuario_sistema,
            titulo="Lectura confirmada",
            mensaje=f"Lectura confirmada desde estimación para el medidor {medidor.num_medidor}. Consumo: {lectura.consumo_m3} m3. {mensaje_factura}",
            tipo="exito"
        )
        
        # Notificación para el afiliado
        if id_usuario_afiliado:
            mensaje_afiliado = f"Se confirmó la lectura de {lectura.consumo_m3} m3 para tu medidor {medidor.num_medidor}."
            
            if factura_generada:
                mensaje_afiliado += f" Factura {factura_generada.num_factura} generada por ${factura_generada.total}."
            
            registrar_notificacion(
                db=db,
                id_usuario=id_usuario_afiliado,
                titulo="Lectura confirmada",
                mensaje=mensaje_afiliado,
                tipo="info"
            )
        
        # Preparar respuesta
        response_data = lectura_to_response(lectura)
        
        if factura_generada:
            # Obtener tarifa aplicada
            tarifa_aplicada = db.query(Tarifa).filter(
                Tarifa.id_tarifa == factura_generada.id_tarifa
            ).first()
            
            response_data['factura_generada'] = {
                'id_factura': factura_generada.id_factura,
                'num_factura': factura_generada.num_factura,
                'total': float(factura_generada.total),
                'tarifa_aplicada': tarifa_aplicada.tipo_tarifa if tarifa_aplicada else "N/A",
                'periodo': factura_generada.periodo,
                'mensaje': mensaje_factura
            }
        else:
            response_data['factura_generada'] = None
            response_data['mensaje_factura'] = mensaje_factura if mensaje_factura else "Factura no generada"
        
        return response_data
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Error confirmando lectura: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al confirmar lectura: {str(e)}"
        )


# ========================================
# ENDPOINT PARA CONFIRMAR TODAS LAS LECTURAS ESTIMADAS
# ========================================

@router.patch("/confirmar-todas-estimadas", response_model=dict)
def confirmar_todas_lecturas_estimadas(
    mes: int = Query(..., ge=1, le=12, description="Mes del periodo"),
    anio: int = Query(..., ge=2020, le=2100, description="Año del periodo"),
    generar_facturas: bool = Query(True, description="Generar facturas automáticamente"),
    tipo_descuento: str = Query('ninguno', description="Tipo: ninguno/porcentaje/valor"),
    valor_descuento: float = Query(0.0, ge=0, description="Valor del descuento"),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Confirma todas las lecturas estimadas de un periodo específico.
    Convierte las lecturas estimadas en lecturas reales y genera las facturas.
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "lecturas", "actualizar")
    
    try:
        from datetime import date
        import calendar
        
        # Calcular rango de fechas del periodo
        periodo_consumo = construir_periodo_consumo(anio, mes)
        fecha_inicio = date(anio, mes, 1)
        ultimo_dia = calendar.monthrange(anio, mes)[1]
        fecha_fin = date(anio, mes, ultimo_dia)
        
        print(f"🔍 Buscando lecturas estimadas entre {fecha_inicio} y {fecha_fin}")
        
        # Obtener todas las lecturas estimadas del periodo
        lecturas_estimadas = db.query(Lectura).filter(
            Lectura.periodo_consumo == periodo_consumo,
            Lectura.es_estimada == True,
            Lectura.activo == True
        ).all()
        
        print(f"✅ Encontradas {len(lecturas_estimadas)} lecturas estimadas")
        
        if not lecturas_estimadas:
            return {
                "success": True,
                "mensaje": "No hay lecturas estimadas en este periodo",
                "periodo": f"{mes:02d}/{anio}",
                "lecturas_confirmadas": 0,
                "lecturas_fallidas": 0,
                "facturas_generadas": 0,
                "facturas_fallidas": 0,
                "detalles": [],
                "fallidas": []
            }
        
        confirmadas = []
        fallidas = []
        facturas_generadas_count = 0
        facturas_fallidas_count = 0
        
        # Procesar cada lectura estimada
        for lectura in lecturas_estimadas:
            try:
                medidor = lectura.medidor
                
                # Validar que tenga medidor
                if not medidor:
                    fallidas.append({
                        "id_lectura": lectura.id_lectura,
                        "medidor": "N/A",
                        "razon": "Medidor no encontrado"
                    })
                    continue
                
                # Información del afiliado
                afiliado = medidor.usuario_afiliado if medidor else None
                usuario_afiliado = afiliado.usuario_sistema if afiliado else None
                
                if usuario_afiliado:
                    nombre_afiliado = f"{usuario_afiliado.nombres} {usuario_afiliado.apellidos}"
                    id_usuario_afiliado = usuario_afiliado.id_usuario_sistema
                else:
                    nombre_afiliado = "Sin afiliado"
                    id_usuario_afiliado = None
                
                # Convertir a lectura real
                lectura.es_estimada = False
                lectura.id_lector = current_user.id_usuario_sistema
                
                # Actualizar observación
                if lectura.observacion:
                    lectura.observacion = "Lectura confirmada desde estimación"
                else:
                    lectura.observacion = "Lectura confirmada desde estimación"
                
                db.flush()  # Guardar cambios de lectura
                
                # ✅ GENERAR FACTURA PARA ESTA LECTURA
                factura_generada = None
                mensaje_factura = "Sin factura"
                
                if generar_facturas:
                    exito, mensaje, factura_generada = generar_factura_desde_lectura(
                        db=db,
                        lectura=lectura,
                        tipo_descuento=tipo_descuento,
                        valor_descuento=valor_descuento,
                        aplicar_servicios=True,
                        aplicar_multas=True
                    )
                    
                    if exito and factura_generada:
                        facturas_generadas_count += 1
                        mensaje_factura = f"Factura {factura_generada.num_factura}: ${factura_generada.total}"
                        
                        # Notificar al afiliado sobre la factura
                        if id_usuario_afiliado:
                            registrar_notificacion(
                                db=db,
                                id_usuario=id_usuario_afiliado,
                                titulo="Lectura confirmada",
                                mensaje=f"Se confirmó la lectura de {lectura.consumo_m3} m3 para tu medidor {medidor.num_medidor}. Factura {factura_generada.num_factura} generada por ${factura_generada.total}.",
                                tipo="info"
                            )
                    else:
                        facturas_fallidas_count += 1
                        mensaje_factura = f"Error: {mensaje}"
                
                confirmadas.append({
                    "id_lectura": lectura.id_lectura,
                    "medidor": medidor.num_medidor,
                    "nombre_afiliado": nombre_afiliado,
                    "lectura_anterior": lectura.lectura_anterior,
                    "lectura_confirmada": lectura.lectura_actual,
                    "consumo": lectura.consumo_m3,
                    "factura": mensaje_factura
                })
                
                print(f"✅ Lectura {lectura.id_lectura} confirmada: {medidor.num_medidor} - {nombre_afiliado} - {mensaje_factura}")
                
            except Exception as e:
                print(f"❌ Error procesando lectura {lectura.id_lectura}: {e}")
                import traceback
                traceback.print_exc()
                
                # Intentar obtener el número de medidor de forma segura
                num_medidor = "N/A"
                try:
                    if lectura.medidor:
                        num_medidor = lectura.medidor.num_medidor
                except:
                    pass
                
                fallidas.append({
                    "id_lectura": lectura.id_lectura,
                    "medidor": num_medidor,
                    "razon": str(e)
                })
        
        # Guardar cambios
        if confirmadas:
            db.commit()
            print(f"💾 Guardadas {len(confirmadas)} lecturas confirmadas")
            print(f"📄 Generadas {facturas_generadas_count} facturas")
            
            # Auditoría
            registrar_auditoria(
                db=db,
                accion="CONFIRMAR_TODAS_ESTIMADAS",
                descripcion=f"Confirmadas {len(confirmadas)} lecturas estimadas del periodo {mes:02d}/{anio}. Facturas generadas: {facturas_generadas_count}",
                id_usuario=current_user.id_usuario_sistema
            )
            
            # Notificación al usuario que ejecutó la acción
            registrar_notificacion(
                db=db,
                id_usuario=current_user.id_usuario_sistema,
                titulo="Confirmación masiva completada",
                mensaje=f"Se confirmaron {len(confirmadas)} lecturas y se generaron {facturas_generadas_count} facturas para el periodo {mes:02d}/{anio}",
                tipo="exito"
            )
        else:
            db.rollback()
            print("⚠️ No hay lecturas para confirmar, todas fallaron")
        
        # Construir mensaje de respuesta
        mensaje = f"Se confirmaron {len(confirmadas)} de {len(lecturas_estimadas)} lecturas"
        if generar_facturas:
            mensaje += f" y se generaron {facturas_generadas_count} facturas"
        if fallidas:
            mensaje += f" ({len(fallidas)} lecturas fallidas)"
        if facturas_fallidas_count > 0:
            mensaje += f" ({facturas_fallidas_count} facturas fallidas)"
        
        return {
            "success": True,
            "mensaje": mensaje,
            "periodo": f"{mes:02d}/{anio}",
            "lecturas_confirmadas": len(confirmadas),
            "lecturas_fallidas": len(fallidas),
            "facturas_generadas": facturas_generadas_count,
            "facturas_fallidas": facturas_fallidas_count,
            "detalles": confirmadas[:50],  # Primeras 50 para no sobrecargar
            "fallidas": fallidas[:10]  # Primeras 10 fallidas
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Error confirmando lecturas masivamente: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al confirmar lecturas: {str(e)}"
        )
