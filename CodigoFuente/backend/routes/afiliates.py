# routes/affiliates.py
from fastapi import APIRouter, Depends, HTTPException, status, Query
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from sqlalchemy import String, cast

from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from psycopg2.errors import ForeignKeyViolation, UniqueViolation, NotNullViolation
from typing import List, Optional
from datetime import date, datetime
from models.affiliate import UsuarioAfiliado
from models.user import UsuarioSistema
from models.sector import Sector
from models.role import RolAccion
from models.meter import Medidor
from schemas.affiliate import (
    AffiliateCreate, 
    AffiliateUpdate, 
    AffiliateResponse,
    AffiliateWithUserInfo,
    AffiliateBulkCreate,
    AffiliateBulkCreateRequest,
    AffiliateBulkResponse,
    AffiliateBulkResult,
    AffiliateBulkError
)
from schemas.notification import NotificacionCreate
from utils.notifications import registrar_notificacion
from utils.audit_logger import registrar_auditoria
from db.session import SessionLocal
from security.jwt import verify_token
import base64
import io 
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/affiliates", tags=["affiliates"])

def get_db():
    """Dependencia para obtener la sesión de base de datos"""
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ============================================================================
# HELPER: Obtener usuario actual desde el token
# ============================================================================
def get_current_user(payload: dict, db: Session) -> UsuarioSistema:
    """Obtiene el usuario actual desde el payload del JWT"""
    user = db.query(UsuarioSistema).filter(
        UsuarioSistema.usuario == payload["sub"]
    ).first()
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario no encontrado"
        )
    
    return user

# ============================================================================
# HELPER: Verificar permisos de usuario
# ============================================================================
def check_permission(user: UsuarioSistema, db: Session, module: str, action: str = None) -> bool:
    """
    Verifica si el usuario tiene permiso para una acción.
    """
    module = module.lower().strip()
    action = action.lower().strip() if action else None

    permisos = db.query(RolAccion).filter(
        RolAccion.id_rol == user.id_rol,
        RolAccion.activo == True
    ).all()

    acciones_usuario = set()

    for permiso in permisos:
        if not permiso.nombre_accion:
            continue

        perm_module = permiso.nombre_accion.lower().strip()
        perm_action = (permiso.tipo_accion or '').lower().strip()

        if perm_module != module:
            continue

        if perm_action in ['crud', 'operaciones crud']:
            return True

        acciones_usuario.add(perm_action)

    if action is None:
        return bool(acciones_usuario)

    if action in ['leer', 'lectura']:
        if any(a in acciones_usuario for a in ['lectura', 'leer', 'crear', 'actualizar', 'eliminar']):
            return True

    return action in acciones_usuario


def require_permission(user: UsuarioSistema, db: Session, module: str, action: str = None):
    """Verifica permiso y lanza excepción si no lo tiene"""
    if not check_permission(user, db, module, action):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"No tienes permisos para {action or 'acceder a'} {module}"
        )
# ============================================================================
# HELPER: Procesar foto
# ============================================================================
def process_user_photo(foto_bytes):
    """Procesa la foto del usuario para enviarla al frontend"""
    if foto_bytes:
        try:
            foto_base64 = base64.b64encode(foto_bytes).decode('utf-8')
            return f"data:image/jpeg;base64,{foto_base64}"
        except Exception as e:
            print(f"Error procesando foto: {e}")
            return None
    return None


# ============================================================================
# HELPER: Convertir afiliado a respuesta con información completa
# ============================================================================
def affiliate_to_response(affiliate: UsuarioAfiliado, db: Session) -> dict:
    """Convierte un afiliado con información del usuario y sector"""

    # Obtener información del usuario y sector
    user = affiliate.usuario_sistema
    sector = affiliate.sector

    # Procesar la foto del usuario (si existe)
    foto_url = process_user_photo(user.foto) if user and user.foto else None

    return {
        "id_usuario_afi": affiliate.id_usuario_afi,
        "cod_usuario_afi": affiliate.cod_usuario_afi,
        "fecha_afiliacion": affiliate.fecha_afiliacion.isoformat() if affiliate.fecha_afiliacion else None,
        "id_sector": affiliate.id_sector,
        "id_usuario_sistema": affiliate.id_usuario_sistema,
        "activo": affiliate.activo,

        # Información del usuario del sistema
        "usuario": {
            "id": user.id_usuario_sistema,
            "usuario": user.usuario,
            "nombres": user.nombres,
            "apellidos": user.apellidos,
            "foto": foto_url,  # ✅ Aquí usamos la foto procesada
            "cedula": user.cedula,
            "email": user.email,
            "telefono": user.telefono,
            "direccion": user.direccion,
            "activo": user.activo
        } if user else None,

        # Información del sector
        "sector": {
            "id_sector": sector.id_sector,
            "nombre_sector": sector.nombre_sector,
            "descripcion": sector.descripcion,
            "activo": sector.activo
        } if sector else None,

        # informacion de medidor 
        "medidor": [
            {
                "id_medidor": medidor.id_medidor,
                "num_medidor": medidor.num_medidor,
                "latitud": float(medidor.latitud) if medidor.latitud is not None else None,
                "longitud": float(medidor.longitud) if medidor.longitud is not None else None,
                "altitud": float(medidor.altitud) if medidor.altitud is not None else None,
                "activo": medidor.activo
            }
            for medidor in affiliate.medidores
        ] if affiliate.medidores else []
    }

# ========================================
# LISTAR AFILIADOS
# ========================================
@router.get("/", response_model=List[dict])
def listar_afiliados(
    search: Optional[str] = Query(None, description="Buscar por nombre, cédula, código"),
    id_sector: Optional[int] = Query(None, description="Filtrar por sector"),
    activo: Optional[bool] = Query(None, description="Filtrar por estado activo"),
    skip: int = Query(0, ge=0, description="Número de registros a saltar"),
    limit: int = Query(100, ge=1, le=1000, description="Número máximo de registros"),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Lista todos los afiliados con filtros opcionales
    Requiere permiso: afiliados.lectura o afiliados.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "afiliados", "lectura")
    
    query = db.query(UsuarioAfiliado).join(UsuarioSistema).join(Sector)
    
    # Filtro de búsqueda
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            (UsuarioSistema.nombres.ilike(search_filter)) |
            (UsuarioSistema.apellidos.ilike(search_filter)) |
            (UsuarioSistema.cedula.ilike(search_filter)) |
            cast(UsuarioAfiliado.cod_usuario_afi, String).ilike(search_filter)
        )
    
    # Filtro por sector
    if id_sector:
        query = query.filter(UsuarioAfiliado.id_sector == id_sector)
    
    # Filtro por estado
    if activo is not None:
        query = query.filter(UsuarioAfiliado.activo == activo)
    
    # Ordenar por código de afiliado
    query = query.order_by(UsuarioAfiliado.cod_usuario_afi.desc())
    
    # Paginación
    affiliates = query.offset(skip).limit(limit).all()
    
    return [affiliate_to_response(aff, db) for aff in affiliates]

# ========================================
# OBTENER AFILIADO POR ID
# ========================================
@router.get("/{id_usuario_afi}", response_model=dict)
def obtener_afiliado(
    id_usuario_afi: int,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Obtiene un afiliado específico por ID
    Requiere permiso: afiliados.lectura o afiliados.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "afiliados", "lectura")
    
    affiliate = db.query(UsuarioAfiliado).filter(
        UsuarioAfiliado.id_usuario_afi == id_usuario_afi
    ).first()
    
    if not affiliate:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Afiliado no encontrado"
        )
    

    return affiliate_to_response(affiliate, db)

# ========================================
# LISTAR USUARIOS NO AFILIADOS
# ========================================
@router.get("/available/users", response_model=List[dict])
def listar_usuarios_disponibles(
    search: Optional[str] = Query(None, description="Buscar por nombre o cédula"),
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Lista usuarios del sistema que NO están afiliados
    Requiere permiso: afiliados.lectura o afiliados.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "afiliados", "lectura")
    
    # Obtener IDs de usuarios ya afiliados
    afiliados_ids = db.query(UsuarioAfiliado.id_usuario_sistema).filter(
        UsuarioAfiliado.activo == True
    ).all()
    afiliados_ids = [id[0] for id in afiliados_ids]
    
    # Buscar usuarios que NO estén afiliados y estén activos
    query = db.query(UsuarioSistema).filter(
        UsuarioSistema.activo == True,
        ~UsuarioSistema.id_usuario_sistema.in_(afiliados_ids) if afiliados_ids else True
    )
    
    # Filtro de búsqueda
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            (UsuarioSistema.nombres.ilike(search_filter)) |
            (UsuarioSistema.apellidos.ilike(search_filter)) |
            (UsuarioSistema.cedula.ilike(search_filter))
        )
    
    users = query.order_by(UsuarioSistema.nombres).all()
    
    return [{
        "id_usuario_sistema": user.id_usuario_sistema,
        "usuario": user.usuario,
        "nombres": user.nombres,
        "apellidos": user.apellidos,
        "cedula": user.cedula,
        "email": user.email,
        "telefono": user.telefono,
        "direccion": user.direccion
    } for user in users]

# ========================================
# CREAR AFILIADO
# ========================================
@router.post("/", response_model=dict, status_code=status.HTTP_201_CREATED)
def crear_afiliado(
    affiliate_data: AffiliateCreate,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Crea un nuevo afiliado vinculando un usuario del sistema con un sector
    Requiere permiso: afiliados.crear o afiliados.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "afiliados", "crear")
    
    # Verificar que el usuario del sistema existe
    user = db.query(UsuarioSistema).filter(
        UsuarioSistema.id_usuario_sistema == affiliate_data.id_usuario_sistema
    ).first()
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Usuario del sistema no encontrado"
        )
    
    # Verificar que el sector existe
    sector = db.query(Sector).filter(
        Sector.id_sector == affiliate_data.id_sector
    ).first()
    
    if not sector:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Sector no encontrado"
        )
    
    # Verificar que el usuario no esté ya afiliado
    existe = db.query(UsuarioAfiliado).filter(
        UsuarioAfiliado.id_usuario_sistema == affiliate_data.id_usuario_sistema,
        UsuarioAfiliado.activo == True
    ).first()
    
    if existe:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"El usuario '{user.nombres} {user.apellidos}' ya está afiliado"
        )
    
    # Generar código de afiliado (último código + 1)
    ultimo_codigo = db.query(UsuarioAfiliado.cod_usuario_afi).order_by(
        UsuarioAfiliado.cod_usuario_afi.desc()
    ).first()
    
    nuevo_codigo = (ultimo_codigo[0] + 1) if ultimo_codigo else 1
    
    # Crear nuevo afiliado
    nuevo_afiliado = UsuarioAfiliado(
        cod_usuario_afi=nuevo_codigo,
        fecha_afiliacion=date.today(),
        id_sector=affiliate_data.id_sector,
        id_usuario_sistema=affiliate_data.id_usuario_sistema,
        activo=True
    )
    
    try:
        db.add(nuevo_afiliado)
        db.commit()
        db.refresh(nuevo_afiliado)
        
        # Auditoría
        registrar_auditoria(
            db=db,
            accion="CREATE",
            descripcion=f"Afiliado creado: {user.nombres} {user.apellidos} (Código: {nuevo_codigo}) en sector '{sector.nombre_sector}' por '{payload['sub']}'",
            id_usuario=current_user.id_usuario_sistema
        )
        
        # Notificación
        registrar_notificacion(
            db=db,
            id_usuario=current_user.id_usuario_sistema,
            titulo="Afiliado creado",
            mensaje=f"El usuario '{user.nombres} {user.apellidos}' fue afiliado correctamente con código {nuevo_codigo}.",
            tipo="exito"
        )
        
        return affiliate_to_response(nuevo_afiliado, db)
    
    except IntegrityError as e:
        db.rollback()
        if isinstance(e.orig, UniqueViolation):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El usuario ya está afiliado"
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear afiliado: {str(e)}"
        )
    except Exception as e:
        db.rollback()
        print(f"❌ Error al crear afiliado: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear el afiliado: {str(e)}"
        )

# ========================================
# ACTUALIZAR AFILIADO
# ========================================
@router.put("/{id_usuario_afi}", response_model=dict)
def actualizar_afiliado(
    id_usuario_afi: int,
    affiliate_data: AffiliateUpdate,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Actualiza un afiliado existente (cambiar sector principalmente)
    Requiere permiso: afiliados.actualizar o afiliados.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "afiliados", "actualizar")
    
    # Buscar el afiliado
    affiliate = db.query(UsuarioAfiliado).filter(
        UsuarioAfiliado.id_usuario_afi == id_usuario_afi
    ).first()
    
    if not affiliate:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Afiliado no encontrado"
        )
    
    # Si se cambia el sector, verificar que existe
    if affiliate_data.id_sector and affiliate_data.id_sector != affiliate.id_sector:
        sector = db.query(Sector).filter(
            Sector.id_sector == affiliate_data.id_sector
        ).first()
        
        if not sector:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Sector no encontrado"
            )
    
    # Actualizar campos
    update_data = affiliate_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(affiliate, key, value)
    
    try:
        db.commit()
        db.refresh(affiliate)
        
        # Auditoría
        user = affiliate.usuario_sistema
        registrar_auditoria(
            db=db,
            accion="UPDATE",
            descripcion=f"Afiliado actualizado: {user.nombres} {user.apellidos} (Código: {affiliate.cod_usuario_afi}) por '{payload['sub']}'",
            id_usuario=current_user.id_usuario_sistema
        )
        
        # Notificación
        registrar_notificacion(
            db=db,
            id_usuario=current_user.id_usuario_sistema,
            titulo="Afiliado modificado",
            mensaje=f"El afiliado '{user.nombres} {user.apellidos}' fue modificado correctamente.",
            tipo="info"
        )
        
        return affiliate_to_response(affiliate, db)
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error al actualizar afiliado: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al actualizar el afiliado"
        )

# ========================================
# ELIMINAR AFILIADO
# ========================================
@router.delete("/{id_usuario_afi}", status_code=status.HTTP_200_OK)
def eliminar_afiliado(
    id_usuario_afi: int,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Elimina el afiliado si no tiene relaciones.
    Si tiene relaciones → NO lo elimina, NO lo desactiva, solo notifica.
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "afiliados", "eliminar")
    
    affiliate = db.query(UsuarioAfiliado).filter(
        UsuarioAfiliado.id_usuario_afi == id_usuario_afi
    ).first()
    
    if not affiliate:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Afiliado no encontrado"
        )
    
    user = affiliate.usuario_sistema
    nombre_completo = f"{user.nombres} {user.apellidos}"
    codigo = affiliate.cod_usuario_afi
    
    try:
        # Intentar eliminar físicamente
        db.delete(affiliate)
        db.commit()
        
        # Auditoría
        registrar_auditoria(
            db=db,
            accion="DELETE",
            descripcion=f"Afiliado eliminado: {nombre_completo} (Código: {codigo}) por '{payload['sub']}'",
            id_usuario=current_user.id_usuario_sistema
        )
        
        registrar_notificacion(
            db=db,
            id_usuario=current_user.id_usuario_sistema,
            titulo="Afiliado eliminado",
            mensaje=f"El afiliado '{nombre_completo}' fue eliminado correctamente.",
            tipo="info"
        )
        
        return {
            "success": True,
            "accion": "eliminado",
            "message": f"Afiliado '{nombre_completo}' eliminado correctamente."
        }
    
    except IntegrityError as e:
        db.rollback()

        # ✔️ Si es por FK o NOT NULL → NO borrar, NO desactivar
        if isinstance(e.orig, (ForeignKeyViolation, NotNullViolation)):
            
            registrar_notificacion(
                db=db,
                id_usuario=current_user.id_usuario_sistema,
                titulo="Afiliado no eliminado",
                mensaje=f"El afiliado '{nombre_completo}' no se puede eliminar porque tiene relaciones con otros módulos.",
                tipo="alerta"
            )

            return {
                "success": False,
                "accion": "no_eliminado",
                "message": (
                    f"⚠️ El afiliado '{nombre_completo}' NO se puede eliminar "
                    "porque está relacionado con otros módulos (medidores, facturas, etc.)."
                )
            }

        # Otros errores inesperados
        raise HTTPException(
            status_code=500,
            detail="Error al intentar eliminar el afiliado"
        )

# ========================================
# CAMBIAR ESTADO (ACTIVAR/DESACTIVAR)
# ========================================
@router.patch("/{id_usuario_afi}/toggle-status", response_model=dict)
def toggle_affiliate_status(
    id_usuario_afi: int,
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Activa o desactiva un afiliado
    Requiere permiso: afiliados.actualizar o afiliados.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "afiliados", "actualizar")
    
    affiliate = db.query(UsuarioAfiliado).filter(
        UsuarioAfiliado.id_usuario_afi == id_usuario_afi
    ).first()
    
    if not affiliate:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Afiliado no encontrado"
        )
    
    # Cambiar estado
    affiliate.activo = not affiliate.activo
    estado_texto = "activado" if affiliate.activo else "desactivado"
    
    try:
        db.commit()
        db.refresh(affiliate)
        
        user = affiliate.usuario_sistema
        # Auditoría
        registrar_auditoria(
            db=db,
            accion="UPDATE",
            descripcion=f"Afiliado {estado_texto}: {user.nombres} {user.apellidos} (Código: {affiliate.cod_usuario_afi}) por '{payload['sub']}'",
            id_usuario=current_user.id_usuario_sistema
        )
        
        return affiliate_to_response(affiliate, db)
    
    except Exception as e:
        db.rollback()
        print(f"❌ Error al cambiar estado: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al cambiar el estado del afiliado"
        )

# ========================================
# ESTADÍSTICAS
# ========================================
@router.get("/stats/count")
def obtener_estadisticas_afiliados(
    db: Session = Depends(get_db),
    payload: dict = Depends(verify_token)
):
    """
    Obtiene estadísticas de afiliados
    Requiere permiso: afiliados.lectura o afiliados.crud
    """
    current_user = get_current_user(payload, db)
    require_permission(current_user, db, "afiliados", "lectura")
    
    total = db.query(UsuarioAfiliado).count()
    activos = db.query(UsuarioAfiliado).filter(UsuarioAfiliado.activo == True).count()
    inactivos = db.query(UsuarioAfiliado).filter(UsuarioAfiliado.activo == False).count()
    
    # Afiliados por sector
    por_sector = db.query(
        Sector.nombre_sector,
        db.func.count(UsuarioAfiliado.id_usuario_afi)
    ).join(
        UsuarioAfiliado, Sector.id_sector == UsuarioAfiliado.id_sector
    ).filter(
        UsuarioAfiliado.activo == True
    ).group_by(
        Sector.nombre_sector
    ).all()
    
    return {
        "total": total,
        "activos": activos,
        "inactivos": inactivos,
        "por_sector": [{"sector": s[0], "cantidad": s[1]} for s in por_sector]
    }
# ========================================
# DESCARGAR PLANTILLA EXCEL 
# ========================================
@router.get("/template/download")
def download_template(
    payload: dict = Depends(verify_token),
    db: Session = Depends(get_db)
):
    """
    Descarga una plantilla de Excel con:
    - Usuarios disponibles (no afiliados)
    - Sectores disponibles
    - Formato correcto para carga masiva
    """
    current_user = get_current_user(payload, db)
    
    try:
        # Obtener usuarios NO afiliados
        afiliados_ids = db.query(UsuarioAfiliado.id_usuario_sistema).filter(
            UsuarioAfiliado.activo == True
        ).all()
        afiliados_ids = [id[0] for id in afiliados_ids]

        # Usuarios disponibles = NO afiliados
        usuarios_disponibles = db.query(UsuarioSistema).filter(
            UsuarioSistema.activo == True,
            ~UsuarioSistema.id_usuario_sistema.in_(afiliados_ids) if afiliados_ids else True
        ).order_by(UsuarioSistema.nombres).all()

        
        # Obtener sectores
        sectores = db.query(Sector).filter(Sector.activo == True).all()
        
        print(f"📊 Generando plantilla con {len(usuarios_disponibles)} usuarios y {len(sectores)} sectores")
        
        # Crear libro de Excel
        wb = Workbook()
        
       # ===============================
        # HOJA 1: PLANTILLA PARA LLENAR
        # ===============================
        ws_plantilla = wb.active
        ws_plantilla.title = "Plantilla Afiliados"

        # Encabezados
        headers = [
            "id_usuario_sistema",
            "nombres",
            "apellidos",
            "id_sector",      # usuario llena
            "num_medidor",    # usuario llena
            "latitud",        # usuario llena
            "longitud",       # usuario llena
            "altitud"         # usuario llena
        ]

        # Estilo encabezado
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws_plantilla.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ANCHOS DE COLUMNA
        column_widths = [20, 25, 25, 15, 18, 15, 15, 15]
        for col, width in zip("ABCDEFGH", column_widths):
            ws_plantilla.column_dimensions[col].width = width


        # =======================================================
        # AGREGAR TODAS LAS FILAS AUTOMÁTICAMENTE (USUARIOS)
        # =======================================================
        for row_num, usuario in enumerate(usuarios_disponibles, 2):
            ws_plantilla.cell(row=row_num, column=1, value=usuario.id_usuario_sistema)
            ws_plantilla.cell(row=row_num, column=2, value=usuario.nombres)
            ws_plantilla.cell(row=row_num, column=3, value=usuario.apellidos)

            ws_plantilla.cell(row=row_num, column=4, value="")    # id_sector -> usuario llena
            ws_plantilla.cell(row=row_num, column=5, value="")    # num_medidor
            ws_plantilla.cell(row=row_num, column=6, value="")    # latitud
            ws_plantilla.cell(row=row_num, column=7, value="")    # longitud
            ws_plantilla.cell(row=row_num, column=8, value="")    # altitud

                
        # ===============================
        # HOJA 2: USUARIOS DISPONIBLES
        # ===============================
        ws_usuarios = wb.create_sheet("Usuarios Disponibles")
        
        # Encabezados
        headers_usuarios = ["id_usuario_sistema", "nombres", "apellidos", "cedula", "email"]
        
        for col_num, header in enumerate(headers_usuarios, 1):
            cell = ws_usuarios.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Datos de usuarios
        if usuarios_disponibles:
            for row_num, usuario in enumerate(usuarios_disponibles, 2):
                ws_usuarios.cell(row=row_num, column=1, value=usuario.id_usuario_sistema)
                ws_usuarios.cell(row=row_num, column=2, value=usuario.nombres)
                ws_usuarios.cell(row=row_num, column=3, value=usuario.apellidos)
                ws_usuarios.cell(row=row_num, column=4, value=usuario.cedula)
                ws_usuarios.cell(row=row_num, column=5, value=usuario.email)
        else:
            # Mensaje si no hay usuarios
            ws_usuarios.cell(row=2, column=1, value="No hay usuarios disponibles")
            ws_usuarios.merge_cells('A2:E2')
        
        # Ajustar anchos
        ws_usuarios.column_dimensions['A'].width = 20
        ws_usuarios.column_dimensions['B'].width = 25
        ws_usuarios.column_dimensions['C'].width = 25
        ws_usuarios.column_dimensions['D'].width = 15
        ws_usuarios.column_dimensions['E'].width = 30
        
        # ===============================
        # HOJA 3: SECTORES DISPONIBLES
        # ===============================
        ws_sectores = wb.create_sheet("Sectores Disponibles")
        
        # Encabezados
        headers_sectores = ["id_sector", "nombre_sector"]
        
        for col_num, header in enumerate(headers_sectores, 1):
            cell = ws_sectores.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Datos de sectores
        if sectores:
            for row_num, sector in enumerate(sectores, 2):
                ws_sectores.cell(row=row_num, column=1, value=sector.id_sector)
                ws_sectores.cell(row=row_num, column=2, value=sector.nombre_sector)
        else:
            ws_sectores.cell(row=2, column=1, value="No hay sectores disponibles")
            ws_sectores.merge_cells('A2:B2')
        
        # Ajustar anchos
        ws_sectores.column_dimensions['A'].width = 15
        ws_sectores.column_dimensions['B'].width = 30
        
        # ===============================
        # HOJA 4: INSTRUCCIONES
        # ===============================
        ws_instrucciones = wb.create_sheet("Instrucciones")
        
        instrucciones = [
            ["📋 INSTRUCCIONES PARA CARGA MASIVA DE AFILIADOS Y MEDIDORES"],
            [""],
            ["1️⃣ USO DE LA PLANTILLA:"],
            ["   • Complete SOLO la hoja 'Plantilla Afiliados'"],
            ["   • NO modifique las hojas 'Usuarios Disponibles' ni 'Sectores Disponibles'"],
            ["   • Borre las filas de ejemplo antes de agregar sus datos"],
            [""],
            ["2️⃣ COLUMNAS REQUERIDAS:"],
            ["   • id_usuario_sistema: ID del usuario (consultar hoja 'Usuarios Disponibles')"],
            ["   • id_sector: ID del sector (consultar hoja 'Sectores Disponibles')"],
            ["   • num_medidor: Número único del medidor (ej: MED-001)"],
            ["   • latitud: Coordenada (opcional, formato: -1.234567)"],
            ["   • longitud: Coordenada (opcional, formato: -78.123456)"],
            ["   • altitud: Altura en metros (opcional, formato: 2850.00)"],
            [""],
            ["3️⃣ VALIDACIONES:"],
            ["   • El usuario NO debe estar ya afiliado"],
            ["   • El número de medidor debe ser único"],
            ["   • Máximo 100 afiliados por carga"],
            [""],
            ["4️⃣ PROCESO AUTOMÁTICO:"],
            ["   • Se generará un código de afiliado único automáticamente"],
            ["   • Se creará el afiliado con fecha actual"],
            ["   • Se creará el medidor asociado al afiliado"],
            ["   • Estado por defecto: Activo"],
            [""],
            ["5️⃣ DESPUÉS DE COMPLETAR:"],
            ["   • Guarde el archivo Excel"],
            ["   • Súbalo en el sistema usando el botón 'Crear desde Excel'"],
            ["   • El sistema validará y creará los registros"],
            ["   • Recibirá un reporte de exitosos y fallidos"],
            [""],
            [f"📅 Plantilla generada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
            [f"👤 Usuario: {current_user.nombres} {current_user.apellidos}"],
        ]
        
        for row_num, fila in enumerate(instrucciones, 1):
            cell = ws_instrucciones.cell(row=row_num, column=1, value=fila[0])
            if row_num == 1:
                cell.font = Font(size=14, bold=True, color="4472C4")
            elif any(emoji in str(fila[0]) for emoji in ["1️⃣", "2️⃣", "3️⃣", "4️⃣", "5️⃣"]):
                cell.font = Font(size=12, bold=True)
        
        ws_instrucciones.column_dimensions['A'].width = 80
        
        # ===============================
        # GUARDAR Y RETORNAR - CORREGIDO
        # ===============================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)  # ✅ Importante: volver al inicio del buffer
        
        # Obtener el contenido como bytes
        excel_data = output.getvalue()
        output.close()
        
        print(f"✅ Excel generado correctamente: {len(excel_data)} bytes")
        
        # Registrar auditoría
        registrar_auditoria(
            db=db,
            accion="DOWNLOAD_TEMPLATE",
            descripcion=f"Plantilla de afiliados descargada por '{current_user.usuario}'",
            id_usuario=current_user.id_usuario_sistema
        )
        
        # Nombre del archivo
        filename = f"plantilla_afiliados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # ✅ CREAR NUEVO BytesIO CON LOS DATOS
        final_output = io.BytesIO(excel_data)
        
        # ✅ RETORNAR CON StreamingResponse
        return StreamingResponse(
            final_output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Length": str(len(excel_data))
            }
        )
        
    except Exception as e:
        print(f"❌ Error generando plantilla: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al generar la plantilla: {str(e)}"
        )
    
# ========================================
# CREAR AFILIADOS MASIVAMENTE DESDE EXCEL
# ========================================
@router.post("/bulk", response_model=AffiliateBulkResponse, status_code=status.HTTP_201_CREATED)
def create_affiliates_bulk(
    request: AffiliateBulkCreateRequest,
    payload: dict = Depends(verify_token),
    db: Session = Depends(get_db)
):
    """
    Crea múltiples afiliados con sus medidores desde Excel.
    
    ✅ Por cada fila se crea:
       1. Un registro en t_usuario_afiliado
       2. Un registro en t_medidor asociado
    
    ✅ Generación automática:
       - cod_usuario_afi: Se genera secuencialmente
       - fecha_afiliacion: Fecha actual
       - activo: True
    """
    current_user = get_current_user(payload, db)
    
    exitosos = []
    fallidos = []
    
    print(f"\n{'='*60}")
    print(f"🚀 INICIANDO CARGA MASIVA DE {len(request.affiliates)} AFILIADOS + MEDIDORES")
    print(f"{'='*60}\n")
    
    # Obtener el último código de afiliado
    ultimo_afiliado = db.query(UsuarioAfiliado).order_by(
        UsuarioAfiliado.cod_usuario_afi.desc()
    ).first()
    
    siguiente_codigo = (ultimo_afiliado.cod_usuario_afi + 1) if ultimo_afiliado else 1
    
    # ===============================
    # 🔄 Procesar cada afiliado
    # ===============================
    for index, affiliate_data in enumerate(request.affiliates, start=1):
        fila_numero = index
        
        try:
            print(f"📝 Procesando fila {fila_numero}")
            
            # ===============================
            # 1️⃣ Validar que el usuario existe y no está afiliado
            # ===============================
            usuario = db.query(UsuarioSistema).filter(
                UsuarioSistema.id_usuario_sistema == affiliate_data.id_usuario_sistema
            ).first()
            
            if not usuario:
                raise ValueError("Usuario no encontrado")
            
            afiliado_existente = db.query(UsuarioAfiliado).filter(
                UsuarioAfiliado.id_usuario_sistema == affiliate_data.id_usuario_sistema
            ).first()
            
            if afiliado_existente:
                raise ValueError("El usuario ya está afiliado")
            
            # ===============================
            # 2️⃣ Validar que el sector existe
            # ===============================
            sector = db.query(Sector).filter(
                Sector.id_sector == affiliate_data.id_sector
            ).first()
            
            if not sector:
                raise ValueError("Sector no encontrado")
            
            # ===============================
            # 3️⃣ Validar que el número de medidor no existe
            # ===============================
            medidor_existente = db.query(Medidor).filter(
                Medidor.num_medidor == affiliate_data.num_medidor
            ).first()
            
            if medidor_existente:
                raise ValueError(f"El medidor '{affiliate_data.num_medidor}' ya existe")
            
            # ===============================
            # 4️⃣ Crear afiliado
            # ===============================
            nuevo_afiliado = UsuarioAfiliado(
                id_usuario_sistema=affiliate_data.id_usuario_sistema,
                id_sector=affiliate_data.id_sector,
                cod_usuario_afi=siguiente_codigo,
                fecha_afiliacion=date.today(),
                activo=True
            )
            
            db.add(nuevo_afiliado)
            db.flush()  # Obtener el ID generado
            
            print(f"   ✅ Afiliado creado: cod={siguiente_codigo}, id={nuevo_afiliado.id_usuario_afi}")
            
            # ===============================
            # 5️⃣ Crear medidor asociado
            # ===============================
            nuevo_medidor = Medidor(
                num_medidor=affiliate_data.num_medidor,
                latitud=affiliate_data.latitud,
                longitud=affiliate_data.longitud,
                altitud=affiliate_data.altitud,
                id_usuario_afi=nuevo_afiliado.id_usuario_afi,
                id_sector=affiliate_data.id_sector,
                activo=True
            )
            
            db.add(nuevo_medidor)
            db.flush()
            
            print(f"   ✅ Medidor creado: {affiliate_data.num_medidor}, id={nuevo_medidor.id_medidor}")
            
            # Agregar a exitosos
            exitosos.append(AffiliateBulkResult(
                fila=fila_numero,
                cod_usuario_afi=siguiente_codigo,
                nombre_usuario=f"{usuario.nombres} {usuario.apellidos}",
                cedula=usuario.cedula,
                sector=sector.nombre_sector,
                num_medidor=affiliate_data.num_medidor,
                id_usuario_afi=nuevo_afiliado.id_usuario_afi,
                id_medidor=nuevo_medidor.id_medidor
            ))
            
            # Incrementar código para el siguiente
            siguiente_codigo += 1
            
        except ValueError as ve:
            print(f"   ❌ Error de validación: {str(ve)}")
            fallidos.append(AffiliateBulkError(
                fila=fila_numero,
                id_usuario_sistema=affiliate_data.id_usuario_sistema,
                num_medidor=affiliate_data.num_medidor,
                error=str(ve)
            ))
            
        except IntegrityError as ie:
            db.rollback()
            print(f"   ❌ Error de integridad: {str(ie)}")
            
            error_msg = "Error de base de datos"
            if "usuario_sistema" in str(ie).lower():
                error_msg = "Usuario ya afiliado"
            elif "medidor" in str(ie).lower():
                error_msg = "Medidor duplicado"
            
            fallidos.append(AffiliateBulkError(
                fila=fila_numero,
                id_usuario_sistema=affiliate_data.id_usuario_sistema,
                num_medidor=affiliate_data.num_medidor,
                error=error_msg
            ))
            
        except Exception as e:
            db.rollback()
            print(f"   ❌ Error inesperado: {str(e)}")
            fallidos.append(AffiliateBulkError(
                fila=fila_numero,
                id_usuario_sistema=affiliate_data.id_usuario_sistema if hasattr(affiliate_data, 'id_usuario_sistema') else None,
                num_medidor=affiliate_data.num_medidor if hasattr(affiliate_data, 'num_medidor') else None,
                error=f"Error: {str(e)}"
            ))
    
    # ===============================
    # 6️⃣ Guardar todos los cambios
    # ===============================
    try:
        db.commit()
        print(f"\n✅ Commit exitoso - {len(exitosos)} afiliados y medidores guardados")
        
        # Registrar auditoría
        registrar_auditoria(
            db=db,
            accion="BULK_CREATE_AFFILIATES",
            descripcion=f"Carga masiva: {len(exitosos)} afiliados+medidores creados por '{payload['sub']}'",
            id_usuario=current_user.id_usuario_sistema
        )
        
        # Notificación
        registrar_notificacion(
            db=db,
            id_usuario=current_user.id_usuario_sistema,
            titulo="Carga masiva completada",
            mensaje=f"Se crearon {len(exitosos)} afiliados con medidores. {len(fallidos)} errores.",
            tipo="exito" if len(fallidos) == 0 else "exito"
        )
        
    except Exception as e:
        db.rollback()
        print(f"\n❌ Error al hacer commit: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al guardar afiliados: {str(e)}"
        )
    
    # ===============================
    # 7️⃣ Retornar resumen
    # ===============================
    print(f"\n{'='*60}")
    print(f"📊 RESUMEN DE CARGA MASIVA")
    print(f"{'='*60}")
    print(f"✅ Exitosos: {len(exitosos)}")
    print(f"❌ Fallidos: {len(fallidos)}")
    print(f"📝 Total procesados: {len(request.affiliates)}")
    print(f"{'='*60}\n")
    
    return AffiliateBulkResponse(
        exitosos=exitosos,
        fallidos=fallidos,
        total_procesados=len(request.affiliates),
        total_exitosos=len(exitosos),
        total_fallidos=len(fallidos)
    )